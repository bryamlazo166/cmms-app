"""Reporte Gerencial Excel — resumen ejecutivo listo para presentar.

Genera un workbook con KPIs calculados, graficos nativos de Excel
(tendencia, pareto, distribucion) y el detalle de OTs con la cronologia
completa de tiempos (solicitud -> programada -> inicio real -> fin real).

Usa el mismo payload que el endpoint /api/reports/executive, por lo que
los numeros del Excel siempre coinciden con lo que se ve en pantalla.
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Paleta alineada al theme Industrial_Dark del proyecto Power BI
NAVY = '1F3864'
CYAN = '0A84FF'
GREEN = '1D6B35'

HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
TITLE_FONT = Font(color=NAVY, bold=True, size=16)
SUBTITLE_FONT = Font(color='555555', size=10)
KPI_LABEL_FONT = Font(bold=True, size=11)
KPI_VALUE_FONT = Font(color=NAVY, bold=True, size=12)
ZEBRA_FILL = PatternFill('solid', fgColor='F2F6FC')


def _write_header_row(ws, row, headers, widths=None):
    """Escribe la fila de encabezado con estilo corporativo."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if widths:
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width


def _write_rows(ws, start_row, rows, number_formats=None):
    """Escribe filas de datos con zebra striping. Devuelve la ultima fila usada."""
    r = start_row - 1
    for i, row_vals in enumerate(rows):
        r = start_row + i
        for col, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL
            if number_formats and col in number_formats and isinstance(val, (int, float)):
                cell.number_format = number_formats[col]
    return r


def _sheet_resumen(wb, payload):
    summary = payload.get('summary', {})
    meta = payload.get('meta', {})

    ws = wb.active
    ws.title = 'Resumen'
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    ws['A1'] = 'REPORTE GERENCIAL DE MANTENIMIENTO'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:E2')
    ws['A2'] = (f"Periodo: {meta.get('start_date', '-')} a {meta.get('end_date', '-')} "
                f"({meta.get('window_days', '-')} dias) | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws['A2'].font = SUBTITLE_FONT

    kpis = [
        ('Cumplimiento Programado (%)', summary.get('compliance_percent', 0), 'OTs cerradas / OTs programadas en el periodo'),
        ('OTs Programadas', summary.get('planned_total', 0), 'Con fecha programada dentro del periodo'),
        ('OTs Cerradas (del plan)', summary.get('planned_closed', 0), ''),
        ('Total OTs ejecutadas', summary.get('total_ots', 0), 'Eventos con fecha real/programada en el periodo'),
        ('Preventivos', summary.get('preventive_count', 0), ''),
        ('Correctivos', summary.get('corrective_count', 0), ''),
        ('Disponibilidad (%)', summary.get('availability', 0), 'Basada en horas de paro de correctivos'),
        ('Horas de Paro', summary.get('downtime_hours', 0), 'Indisponibilidad acumulada'),
        ('MTBF (h)', summary.get('mtbf', 0), 'Tiempo medio entre fallas'),
        ('MTTR (h)', summary.get('mttr', 0), 'Tiempo medio de reparacion'),
        ('Costo de Materiales (S/)', summary.get('cost', 0), 'Materiales de almacen consumidos en OTs'),
    ]

    _write_header_row(ws, 4, ['Indicador', 'Valor', 'Observacion'], widths=[34, 16, 52])
    for i, (label, value, note) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = KPI_LABEL_FONT
        vcell = ws.cell(row=r, column=2, value=value)
        vcell.font = KPI_VALUE_FONT
        vcell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
        vcell.alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3, value=note).font = SUBTITLE_FONT
        if i % 2 == 1:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = ZEBRA_FILL

    # Mini-tabla + dona Preventivo vs Correctivo
    base = 5 + len(kpis) + 2
    ws.cell(row=base, column=1, value='Tipo').fill = HEADER_FILL
    ws.cell(row=base, column=1).font = HEADER_FONT
    ws.cell(row=base, column=2, value='Cantidad').fill = HEADER_FILL
    ws.cell(row=base, column=2).font = HEADER_FONT
    ws.cell(row=base + 1, column=1, value='Preventivo')
    ws.cell(row=base + 1, column=2, value=summary.get('preventive_count', 0))
    ws.cell(row=base + 2, column=1, value='Correctivo')
    ws.cell(row=base + 2, column=2, value=summary.get('corrective_count', 0))

    if (summary.get('preventive_count', 0) + summary.get('corrective_count', 0)) > 0:
        pie = PieChart()
        pie.title = 'Distribucion Preventivo vs Correctivo'
        data = Reference(ws, min_col=2, min_row=base, max_row=base + 2)
        cats = Reference(ws, min_col=1, min_row=base + 1, max_row=base + 2)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)
        pie.height, pie.width = 8, 12
        ws.add_chart(pie, f'E{base}')
    return ws


def _sheet_tendencia(wb, payload):
    trend = payload.get('trend', []) or []
    ws = wb.create_sheet('Tendencia Mensual')

    headers = ['Periodo', 'OTs Programadas', 'OTs Cerradas', 'Cumplimiento %',
               'Preventivos', 'Correctivos', 'Horas Paro', 'Disponibilidad %']
    _write_header_row(ws, 1, headers, widths=[12, 16, 14, 15, 12, 12, 12, 16])
    rows = [[t.get('period'), t.get('planned_total'), t.get('planned_closed'),
             t.get('compliance_percent'), t.get('preventive_count'), t.get('corrective_count'),
             t.get('downtime_hours'), t.get('availability')] for t in trend]
    last = _write_rows(ws, 2, rows, number_formats={4: '0.0', 7: '0.00', 8: '0.00'})

    if rows:
        cats = Reference(ws, min_col=1, min_row=2, max_row=last)

        line = LineChart()
        line.title = 'Cumplimiento y Disponibilidad por mes'
        line.y_axis.title = '%'
        line.add_data(Reference(ws, min_col=4, min_row=1, max_row=last), titles_from_data=True)
        line.add_data(Reference(ws, min_col=8, min_row=1, max_row=last), titles_from_data=True)
        line.set_categories(cats)
        line.height, line.width = 9, 16
        ws.add_chart(line, 'J2')

        bar = BarChart()
        bar.type = 'col'
        bar.title = 'Preventivo vs Correctivo por mes'
        bar.y_axis.title = 'OTs'
        bar.add_data(Reference(ws, min_col=5, min_row=1, max_col=6, max_row=last), titles_from_data=True)
        bar.set_categories(cats)
        bar.height, bar.width = 9, 16
        ws.add_chart(bar, 'J22')
    return ws


def _sheet_areas(wb, payload):
    areas = (payload.get('breakdown', {}) or {}).get('areas', []) or []
    ws = wb.create_sheet('Desglose Areas')

    headers = ['Area', 'Programadas', 'Cerradas', 'Cumplimiento %', 'Preventivos',
               'Correctivos', 'Horas Paro', 'Disponibilidad %', 'MTBF (h)', 'MTTR (h)', 'Costo (S/)']
    _write_header_row(ws, 1, headers, widths=[26, 13, 11, 15, 12, 12, 12, 16, 11, 11, 13])
    rows = [[a.get('name'), a.get('planned_total'), a.get('planned_closed'),
             a.get('compliance_percent'), a.get('preventive_count'), a.get('corrective_count'),
             a.get('downtime_hours'), a.get('availability'), a.get('mtbf'), a.get('mttr'),
             a.get('cost')] for a in areas]
    last = _write_rows(ws, 2, rows, number_formats={4: '0.0', 7: '0.00', 8: '0.00',
                                                    9: '0.0', 10: '0.0', 11: '#,##0.00'})
    if rows:
        bar = BarChart()
        bar.type = 'col'
        bar.title = 'Horas de paro por area'
        bar.y_axis.title = 'Horas'
        bar.add_data(Reference(ws, min_col=7, min_row=1, max_row=last), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        bar.height, bar.width = 9, 16
        ws.add_chart(bar, 'M2')
    return ws


def _sheet_pareto(wb, payload):
    causes = payload.get('downtime_causes', []) or []
    ws = wb.create_sheet('Pareto Fallas')

    headers = ['Modo de Falla', 'Eventos', 'Horas Paro', '% del Paro', '% Acumulado', 'Costo (S/)']
    _write_header_row(ws, 1, headers, widths=[34, 10, 12, 12, 13, 13])

    total_down = sum(float(c.get('downtime_hours') or 0) for c in causes) or 1.0
    acum = 0.0
    rows = []
    for c in causes:
        hrs = float(c.get('downtime_hours') or 0)
        pct = hrs / total_down * 100
        acum += pct
        rows.append([c.get('cause'), c.get('count'), hrs, round(pct, 1),
                     round(min(acum, 100.0), 1), c.get('cost')])
    last = _write_rows(ws, 2, rows, number_formats={3: '0.00', 4: '0.0', 5: '0.0', 6: '#,##0.00'})

    if rows:
        bar = BarChart()
        bar.type = 'col'
        bar.title = 'Pareto de indisponibilidad'
        bar.y_axis.title = 'Horas de paro'
        bar.add_data(Reference(ws, min_col=3, min_row=1, max_row=last), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))

        line = LineChart()
        line.add_data(Reference(ws, min_col=5, min_row=1, max_row=last), titles_from_data=True)
        line.y_axis.axId = 200
        line.y_axis.title = '% acumulado'
        bar.y_axis.crosses = 'max'
        bar += line
        bar.height, bar.width = 10, 18
        ws.add_chart(bar, 'H2')
    return ws


def _sheet_eventos(wb, payload):
    events = payload.get('downtime_events', []) or []
    ws = wb.create_sheet('Eventos de Paro')

    headers = ['OT', 'Fecha', 'Area', 'Linea', 'Equipo', 'Modo de Falla',
               'Duracion (h)', 'Costo (S/)', 'Descripcion']
    _write_header_row(ws, 1, headers, widths=[11, 11, 18, 18, 26, 22, 12, 12, 60])
    rows = [[e.get('ot_code'), e.get('date'), e.get('area'), e.get('line'), e.get('equipment'),
             e.get('failure_mode'), e.get('duration_hours'), e.get('cost'),
             e.get('description')] for e in events]
    last = _write_rows(ws, 2, rows, number_formats={7: '0.00', 8: '#,##0.00'})
    if rows:
        ws.auto_filter.ref = f'A1:I{last}'
    ws.freeze_panes = 'A2'
    return ws


def _collect_ot_detail(meta):
    """OTs del periodo con la cronologia completa de tiempos.

    Replica la ventana y filtros del reporte ejecutivo: una OT entra si su
    fecha de evento (fin real > inicio real > programada) cae en el rango.
    """
    from models import (Area, Component, Equipment, Line, MaintenanceNotice,
                        Provider, System, Technician, WorkOrder)

    start = date.fromisoformat(meta['start_date'])
    end = date.fromisoformat(meta['end_date'])
    filters = meta.get('filters') or {}
    f_area, f_line, f_equip = filters.get('area_id'), filters.get('line_id'), filters.get('equipment_id')

    areas = {a.id: a for a in Area.query.all()}
    lines = {l.id: l for l in Line.query.all()}
    equips = {e.id: e for e in Equipment.query.all()}
    systems = {s.id: s for s in System.query.all()}
    comps = {c.id: c for c in Component.query.all()}
    techs = {t.id: t for t in Technician.query.all()}
    provs = {p.id: p for p in Provider.query.all()}
    notices = {n.id: n for n in MaintenanceNotice.query.all()}

    def _d(v):
        try:
            return date.fromisoformat(str(v)[:10])
        except (TypeError, ValueError):
            return None

    def _fmt(v):
        return str(v).replace('T', ' ') if v else '-'

    rows = []
    for o in WorkOrder.query.order_by(WorkOrder.id).all():
        event = _d(o.real_end_date) or _d(o.real_start_date) or _d(o.scheduled_date)
        if not event or event < start or event > end:
            continue

        # Resolver jerarquia (mismo criterio que el reporte ejecutivo)
        eq_id = o.equipment_id
        if not eq_id and o.system_id and o.system_id in systems:
            eq_id = systems[o.system_id].equipment_id
        if not eq_id and o.component_id and o.component_id in comps:
            sys_ = systems.get(comps[o.component_id].system_id)
            if sys_:
                eq_id = sys_.equipment_id
        eq = equips.get(eq_id)
        line_id = o.line_id or (eq.line_id if eq else None)
        ln = lines.get(line_id)
        area_id = o.area_id or (ln.area_id if ln else None)

        if f_equip and eq_id != f_equip:
            continue
        if f_line and line_id != f_line:
            continue
        if f_area and area_id != f_area:
            continue

        nt = notices.get(o.notice_id) if o.notice_id else None
        solicitud = (getattr(nt, 'reported_at', None) or getattr(nt, 'request_date', None)) if nt else None

        tech = None
        try:
            tech = techs.get(int(o.technician_id))
        except (TypeError, ValueError):
            pass

        rows.append([
            o.code or f'OT-{o.id}',
            nt.code if nt else '-',
            o.status,
            o.maintenance_type,
            areas[area_id].name if area_id in areas else '-',
            ln.name if ln else '-',
            eq.name if eq else '-',
            eq.tag if eq else '-',
            _fmt(solicitud),
            o.scheduled_date or '-',
            _fmt(o.real_start_date),
            _fmt(o.real_end_date),
            o.real_duration,
            o.estimated_duration,
            tech.name if tech else (o.technician_id or '-'),
            provs[o.provider_id].name if o.provider_id in provs else '-',
            o.failure_mode or '-',
            o.description or '-',
        ])
    rows.sort(key=lambda r: (str(r[9]), str(r[0])))
    return rows


def _sheet_detalle(wb, payload):
    ws = wb.create_sheet('Detalle OTs')
    headers = ['Codigo OT', 'Aviso', 'Estado', 'Tipo Mtto', 'Area', 'Linea', 'Equipo', 'TAG',
               'F. Solicitud', 'F. Programada', 'Inicio Real', 'Fin Real',
               'Duracion Real (h)', 'Duracion Est (h)', 'Tecnico', 'Proveedor',
               'Modo de Falla', 'Descripcion']
    widths = [11, 11, 11, 11, 18, 18, 24, 12, 17, 14, 17, 17, 15, 14, 18, 18, 20, 60]
    _write_header_row(ws, 1, headers, widths=widths)

    rows = _collect_ot_detail(payload.get('meta', {}))
    last = _write_rows(ws, 2, rows, number_formats={13: '0.00', 14: '0.00'})
    if rows:
        ws.auto_filter.ref = f'A1:R{last}'
    ws.freeze_panes = 'C2'
    return ws


def build_management_workbook(payload):
    """Construye el Excel gerencial completo y devuelve BytesIO."""
    wb = Workbook()
    _sheet_resumen(wb, payload)
    _sheet_tendencia(wb, payload)
    _sheet_areas(wb, payload)
    _sheet_pareto(wb, payload)
    _sheet_eventos(wb, payload)
    _sheet_detalle(wb, payload)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
