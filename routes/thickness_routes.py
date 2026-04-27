"""Rutas para inspección de espesores por ultrasonido (UT)."""
import datetime as dt

from flask import jsonify, request


def register_thickness_routes(
    app,
    db,
    logger,
    ThicknessPoint,
    ThicknessInspection,
    ThicknessReading,
    Equipment,
    MaintenanceNotice=None,
):

    def _today():
        return dt.date.today().isoformat()

    def _calc_status(value, point):
        """Devuelve (status, is_alert, is_critical) según el valor y los umbrales del punto."""
        if value is None:
            return ('NORMAL', False, False)
        if value <= point.scrap_thickness:
            return ('CRITICO', False, True)
        if value <= point.alarm_thickness:
            return ('ALERTA', True, False)
        return ('NORMAL', False, False)

    def _semaphore_for_equipment(equipment_id):
        """Calcula el semáforo de la próxima inspección programada del equipo."""
        last = ThicknessInspection.query.filter_by(equipment_id=equipment_id) \
            .order_by(ThicknessInspection.inspection_date.desc()).first()
        if not last or not last.next_due_date:
            return ('PENDIENTE', None)
        try:
            due = dt.date.fromisoformat(last.next_due_date)
            today = dt.date.today()
            days_left = (due - today).days
            if days_left < 0:
                return ('ROJO', days_left)
            if days_left <= 10:
                return ('AMARILLO', days_left)
            return ('VERDE', days_left)
        except Exception:
            return ('PENDIENTE', None)

    # ── CATALOGO DE PUNTOS ─────────────────────────────────────────────────
    @app.route('/api/thickness/points/<int:equipment_id>', methods=['GET', 'POST'])
    def handle_thickness_points(equipment_id):
        if request.method == 'POST':
            try:
                data = request.json or {}
                pt = ThicknessPoint(
                    equipment_id=equipment_id,
                    component_id=data.get('component_id'),
                    group_name=data.get('group_name', '').upper(),
                    section=data.get('section'),
                    position=data.get('position', '').upper(),
                    nominal_thickness=float(data.get('nominal_thickness', 25.4)),
                    alarm_thickness=float(data.get('alarm_thickness', 10.0)),
                    scrap_thickness=float(data.get('scrap_thickness', 8.0)),
                    order_index=data.get('order_index', 0),
                )
                db.session.add(pt)
                db.session.commit()
                return jsonify(pt.to_dict()), 201
            except Exception as e:
                db.session.rollback()
                return jsonify({"error": str(e)}), 500
        # GET
        points = ThicknessPoint.query.filter_by(
            equipment_id=equipment_id, is_active=True
        ).order_by(ThicknessPoint.group_name, ThicknessPoint.section, ThicknessPoint.order_index).all()
        return jsonify([p.to_dict() for p in points])

    @app.route('/api/thickness/points/<int:point_id>/edit', methods=['PUT'])
    def update_thickness_point(point_id):
        try:
            pt = ThicknessPoint.query.get_or_404(point_id)
            data = request.json or {}
            for key in ('nominal_thickness', 'alarm_thickness', 'scrap_thickness'):
                if key in data:
                    setattr(pt, key, float(data[key]))
            db.session.commit()
            return jsonify(pt.to_dict())
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    # ── INSPECCIONES ───────────────────────────────────────────────────────
    @app.route('/api/thickness/inspections', methods=['GET', 'POST'])
    def handle_thickness_inspections():
        if request.method == 'POST':
            try:
                data = request.json or {}
                equipment_id = int(data['equipment_id'])
                inspection_date = data.get('inspection_date') or _today()
                frequency_days = int(data.get('frequency_days', 60))
                inspector = data.get('inspector_name')
                observations = data.get('observations')
                readings_data = data.get('readings', [])

                # Calcular next_due_date
                try:
                    insp_dt = dt.date.fromisoformat(inspection_date)
                except Exception:
                    insp_dt = dt.date.today()
                next_due = (insp_dt + dt.timedelta(days=frequency_days)).isoformat()

                # Crear inspección
                inspection = ThicknessInspection(
                    equipment_id=equipment_id,
                    inspection_date=inspection_date,
                    next_due_date=next_due,
                    frequency_days=frequency_days,
                    inspector_name=inspector,
                    status='COMPLETA',
                    observations=observations,
                    pdf_url=(data.get('pdf_url') or None),
                )
                db.session.add(inspection)
                db.session.flush()  # obtener id

                total = 0
                criticals = 0
                alerts = 0
                critical_details = []
                # Crear readings
                for r in readings_data:
                    point_id = int(r.get('point_id'))
                    value = r.get('value_mm')
                    if value is None or value == '':
                        continue
                    try:
                        value = float(value)
                    except Exception:
                        continue
                    pt = ThicknessPoint.query.get(point_id)
                    if not pt or pt.equipment_id != equipment_id:
                        continue
                    status, is_alert, is_critical = _calc_status(value, pt)
                    rd = ThicknessReading(
                        inspection_id=inspection.id,
                        point_id=point_id,
                        value_mm=value,
                        is_alert=is_alert,
                        is_critical=is_critical,
                    )
                    db.session.add(rd)
                    # Actualizar punto
                    pt.last_value = value
                    pt.last_date = inspection_date
                    pt.status = status
                    total += 1
                    if is_critical:
                        criticals += 1
                        critical_details.append(f"{pt.group_name} S{pt.section or ''}-{pt.position}: {value}mm (límite {pt.scrap_thickness}mm)")
                    elif is_alert:
                        alerts += 1

                inspection.total_points = total
                inspection.critical_points = criticals
                inspection.alert_points = alerts
                if criticals > 0:
                    inspection.semaphore_status = 'ROJO'
                elif alerts > 0:
                    inspection.semaphore_status = 'AMARILLO'
                else:
                    inspection.semaphore_status = 'VERDE'

                db.session.commit()

                # Generar aviso automático si hay puntos críticos
                if criticals > 0 and MaintenanceNotice:
                    try:
                        eq = Equipment.query.get(equipment_id)
                        eq_name = eq.name if eq else f"Equipo {equipment_id}"
                        desc = (f"Inspección UT detectó {criticals} punto(s) crítico(s) en {eq_name}.\n"
                                f"Detalles:\n- " + "\n- ".join(critical_details[:10]))
                        notice = MaintenanceNotice(
                            description=desc,
                            equipment_id=equipment_id,
                            criticality='Alta',
                            priority='Alta',
                            maintenance_type='Correctivo',
                            failure_category='Estructural',
                            failure_mode='Desgaste',
                            status='Pendiente',
                            scope='PLAN',
                            reporter_name=inspector or 'Sistema UT',
                        )
                        db.session.add(notice)
                        db.session.flush()
                        notice.code = f"AV-{notice.id:04d}"
                        db.session.commit()
                    except Exception as ne:
                        logger.warning(f"thickness: error al crear aviso automático: {ne}")
                        db.session.rollback()

                return jsonify(inspection.to_dict()), 201
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error creando inspección espesores: {e}")
                return jsonify({"error": str(e)}), 500

        # GET — listar inspecciones (filtrar por equipment_id opcional)
        equipment_id = request.args.get('equipment_id', type=int)
        q = ThicknessInspection.query
        if equipment_id:
            q = q.filter_by(equipment_id=equipment_id)
        inspections = q.order_by(ThicknessInspection.inspection_date.desc()).limit(100).all()
        return jsonify([i.to_dict() for i in inspections])

    @app.route('/api/thickness/inspections/<int:inspection_id>/pdf', methods=['PUT'])
    def update_thickness_pdf_url(inspection_id):
        try:
            inspection = ThicknessInspection.query.get_or_404(inspection_id)
            data = request.json or {}
            url = (data.get('pdf_url') or '').strip()
            if not url:
                return jsonify({"error": "pdf_url requerido"}), 400
            inspection.pdf_url = url
            db.session.commit()
            return jsonify(inspection.to_dict())
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    @app.route('/api/thickness/inspections/<int:inspection_id>/edit', methods=['PUT'])
    def edit_thickness_inspection(inspection_id):
        """Editar una inspección existente: actualiza metadata + reemplaza readings."""
        try:
            inspection = ThicknessInspection.query.get_or_404(inspection_id)
            data = request.json or {}
            equipment_id = inspection.equipment_id

            # Actualizar metadata
            if 'inspection_date' in data:
                inspection.inspection_date = data['inspection_date']
                freq = inspection.frequency_days or 60
                try:
                    insp_dt = dt.date.fromisoformat(data['inspection_date'])
                    inspection.next_due_date = (insp_dt + dt.timedelta(days=freq)).isoformat()
                except Exception:
                    pass
            if 'inspector_name' in data:
                inspection.inspector_name = data['inspector_name']
            if 'observations' in data:
                inspection.observations = data['observations']
            if 'pdf_url' in data:
                inspection.pdf_url = data.get('pdf_url') or None

            readings_data = data.get('readings', [])
            if readings_data:
                # Eliminar readings anteriores
                ThicknessReading.query.filter_by(inspection_id=inspection_id).delete()

                total = 0
                criticals = 0
                alerts = 0
                for r in readings_data:
                    point_id = int(r.get('point_id'))
                    value = r.get('value_mm')
                    if value is None or value == '':
                        continue
                    try:
                        value = float(value)
                    except Exception:
                        continue
                    pt = ThicknessPoint.query.get(point_id)
                    if not pt or pt.equipment_id != equipment_id:
                        continue
                    status, is_alert, is_critical = _calc_status(value, pt)
                    rd = ThicknessReading(
                        inspection_id=inspection_id,
                        point_id=point_id,
                        value_mm=value,
                        is_alert=is_alert,
                        is_critical=is_critical,
                    )
                    db.session.add(rd)
                    pt.last_value = value
                    pt.last_date = inspection.inspection_date
                    pt.status = status
                    total += 1
                    if is_critical:
                        criticals += 1
                    elif is_alert:
                        alerts += 1

                inspection.total_points = total
                inspection.critical_points = criticals
                inspection.alert_points = alerts
                if criticals > 0:
                    inspection.semaphore_status = 'ROJO'
                elif alerts > 0:
                    inspection.semaphore_status = 'AMARILLO'
                else:
                    inspection.semaphore_status = 'VERDE'

            db.session.commit()
            return jsonify(inspection.to_dict())
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error editando inspección espesores: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/thickness/inspections/<int:inspection_id>', methods=['GET', 'DELETE'])
    def handle_thickness_inspection_detail(inspection_id):
        inspection = ThicknessInspection.query.get_or_404(inspection_id)
        if request.method == 'DELETE':
            try:
                ThicknessReading.query.filter_by(inspection_id=inspection_id).delete()
                db.session.delete(inspection)
                db.session.commit()
                return jsonify({"message": "Inspección eliminada"})
            except Exception as e:
                db.session.rollback()
                return jsonify({"error": str(e)}), 500
        # GET con readings
        readings = ThicknessReading.query.filter_by(inspection_id=inspection_id).all()
        result = inspection.to_dict()
        result['readings'] = [r.to_dict() for r in readings]
        return jsonify(result)

    # ── DASHBOARD ──────────────────────────────────────────────────────────
    @app.route('/api/thickness/dashboard', methods=['GET'])
    def thickness_dashboard():
        try:
            # Equipos con puntos catalogados
            equipment_ids = db.session.query(ThicknessPoint.equipment_id).distinct().all()
            equipment_ids = [e[0] for e in equipment_ids]
            equipos = []
            for eq_id in equipment_ids:
                eq = Equipment.query.get(eq_id)
                if not eq:
                    continue
                last = ThicknessInspection.query.filter_by(equipment_id=eq_id) \
                    .order_by(ThicknessInspection.inspection_date.desc()).first()
                semaphore, days_left = _semaphore_for_equipment(eq_id)
                point_count = ThicknessPoint.query.filter_by(equipment_id=eq_id, is_active=True).count()
                critical_count = ThicknessPoint.query.filter_by(equipment_id=eq_id, status='CRITICO', is_active=True).count()
                alert_count = ThicknessPoint.query.filter_by(equipment_id=eq_id, status='ALERTA', is_active=True).count()
                equipos.append({
                    "equipment_id": eq_id,
                    "equipment_name": eq.name,
                    "equipment_tag": eq.tag,
                    "last_inspection_date": last.inspection_date if last else None,
                    "next_due_date": last.next_due_date if last else None,
                    "days_left": days_left,
                    "semaphore_status": semaphore,
                    "point_count": point_count,
                    "critical_count": critical_count,
                    "alert_count": alert_count,
                })
            equipos.sort(key=lambda x: (x['equipment_tag'] or ''))
            return jsonify({"equipos": equipos, "total": len(equipos)})
        except Exception as e:
            logger.error(f"thickness_dashboard error: {e}")
            return jsonify({"equipos": [], "total": 0, "error": str(e)}), 200

    # ── HISTORICO POR PUNTO ────────────────────────────────────────────────
    @app.route('/api/thickness/history/<int:point_id>', methods=['GET'])
    def thickness_point_history(point_id):
        readings = db.session.query(ThicknessReading, ThicknessInspection) \
            .join(ThicknessInspection, ThicknessReading.inspection_id == ThicknessInspection.id) \
            .filter(ThicknessReading.point_id == point_id) \
            .order_by(ThicknessInspection.inspection_date.asc()).all()
        return jsonify([{
            "value_mm": r.value_mm,
            "inspection_date": i.inspection_date,
            "is_critical": r.is_critical,
            "is_alert": r.is_alert,
        } for r, i in readings])

    # ── ANÁLISIS PREDICTIVO ────────────────────────────────────────────────
    def _calc_wear_rate(readings_with_dates):
        """Calcula velocidad de desgaste (mm/mes) con regresión lineal simple."""
        if len(readings_with_dates) < 2:
            return None, None
        # Convertir a días desde primera lectura
        try:
            dates = [dt.date.fromisoformat(r['date']) for r in readings_with_dates]
            values = [r['value'] for r in readings_with_dates]
            d0 = dates[0]
            days = [(d - d0).days for d in dates]
            n = len(days)
            if days[-1] == 0:
                return None, None
            # Regresión lineal: y = a + b*x
            sx = sum(days)
            sy = sum(values)
            sxx = sum(d * d for d in days)
            sxy = sum(d * v for d, v in zip(days, values))
            denom = n * sxx - sx * sx
            if denom == 0:
                return None, None
            b = (n * sxy - sx * sy) / denom  # pendiente (mm/día)
            a = (sy - b * sx) / n  # intercepto
            wear_rate_month = abs(b) * 30.44  # mm/mes (positivo = desgaste)
            wear_rate_week = abs(b) * 7
            return {
                'mm_per_day': round(abs(b), 4),
                'mm_per_week': round(wear_rate_week, 3),
                'mm_per_month': round(wear_rate_month, 3),
                'mm_per_year': round(abs(b) * 365.25, 2),
                'slope': round(b, 6),
                'intercept': round(a, 2),
            }, dates[0]
        except Exception:
            return None, None

    @app.route('/api/thickness/analysis/<int:equipment_id>', methods=['GET'])
    def thickness_analysis(equipment_id):
        """Análisis predictivo completo por equipo: desgaste, vida residual, alertas."""
        try:
            points = ThicknessPoint.query.filter_by(
                equipment_id=equipment_id, is_active=True
            ).all()

            eq = Equipment.query.get(equipment_id)
            eq_name = eq.name if eq else f"Equipo {equipment_id}"
            eq_tag = eq.tag if eq else "?"

            analysis = []
            alerts = []

            for pt in points:
                readings = db.session.query(ThicknessReading, ThicknessInspection) \
                    .join(ThicknessInspection, ThicknessReading.inspection_id == ThicknessInspection.id) \
                    .filter(ThicknessReading.point_id == pt.id) \
                    .order_by(ThicknessInspection.inspection_date.asc()).all()

                if not readings:
                    continue

                data_points = [{'date': i.inspection_date, 'value': r.value_mm} for r, i in readings]
                wear, first_date = _calc_wear_rate(data_points)

                last_value = data_points[-1]['value']
                last_date = data_points[-1]['date']
                remaining = last_value - pt.scrap_thickness

                life_months = None
                life_weeks = None
                estimated_replacement = None

                if wear and wear['mm_per_month'] > 0:
                    life_months = round(remaining / wear['mm_per_month'], 1)
                    life_weeks = round(remaining / wear['mm_per_week'], 1) if wear['mm_per_week'] > 0 else None
                    try:
                        last_dt = dt.date.fromisoformat(last_date)
                        replace_dt = last_dt + dt.timedelta(days=int(life_months * 30.44))
                        estimated_replacement = replace_dt.isoformat()
                    except Exception:
                        pass

                entry = {
                    'point_id': pt.id,
                    'group_name': pt.group_name,
                    'section': pt.section,
                    'position': pt.position,
                    'nominal': pt.nominal_thickness,
                    'alarm': pt.alarm_thickness,
                    'scrap': pt.scrap_thickness,
                    'last_value': last_value,
                    'last_date': last_date,
                    'status': pt.status,
                    'readings_count': len(data_points),
                    'readings': data_points,
                    'wear_rate': wear,
                    'remaining_mm': round(remaining, 2),
                    'life_months': life_months,
                    'life_weeks': life_weeks,
                    'estimated_replacement': estimated_replacement,
                }
                analysis.append(entry)

                # Generar alertas si queda poco tiempo
                if life_months is not None:
                    urgency = None
                    if life_months <= 1:
                        urgency = 'CRITICO'
                    elif life_months <= 3:
                        urgency = 'URGENTE'
                    elif life_months <= 6:
                        urgency = 'PLANIFICAR'

                    if urgency:
                        alerts.append({
                            'urgency': urgency,
                            'group_name': pt.group_name,
                            'section': pt.section,
                            'position': pt.position,
                            'last_value': last_value,
                            'scrap': pt.scrap_thickness,
                            'remaining_mm': round(remaining, 2),
                            'wear_mm_month': wear['mm_per_month'],
                            'life_months': life_months,
                            'life_weeks': life_weeks,
                            'estimated_replacement': estimated_replacement,
                            'recommendation': (
                                'REEMPLAZO INMEDIATO — iniciar fabricación URGENTE'
                                if urgency == 'CRITICO' else
                                'Iniciar fabricación AHORA — quedan menos de 3 meses'
                                if urgency == 'URGENTE' else
                                'Programar fabricación en próximas semanas'
                            )
                        })

            # Ordenar alertas por urgencia
            urgency_order = {'CRITICO': 0, 'URGENTE': 1, 'PLANIFICAR': 2}
            alerts.sort(key=lambda a: (urgency_order.get(a['urgency'], 9), a.get('life_months', 999)))

            # Resumen por grupo (componente)
            groups_summary = {}
            for a in analysis:
                g = a['group_name']
                if g not in groups_summary:
                    groups_summary[g] = {
                        'group_name': g,
                        'total_points': 0,
                        'min_value': 999,
                        'max_wear_rate': 0,
                        'min_life_months': 999,
                        'worst_point': None,
                    }
                gs = groups_summary[g]
                gs['total_points'] += 1
                if a['last_value'] < gs['min_value']:
                    gs['min_value'] = a['last_value']
                if a['wear_rate'] and a['wear_rate']['mm_per_month'] > gs['max_wear_rate']:
                    gs['max_wear_rate'] = a['wear_rate']['mm_per_month']
                if a['life_months'] is not None and a['life_months'] < gs['min_life_months']:
                    gs['min_life_months'] = a['life_months']
                    gs['worst_point'] = f"S{a['section'] or ''}-{a['position']}"

            return jsonify({
                'equipment_id': equipment_id,
                'equipment_name': eq_name,
                'equipment_tag': eq_tag,
                'total_points_analyzed': len(analysis),
                'points': analysis,
                'alerts': alerts,
                'groups_summary': list(groups_summary.values()),
            })
        except Exception as e:
            logger.error(f"thickness_analysis error: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": str(e)}), 500

    # ── PLANTILLA EXCEL (descarga + carga) ─────────────────────────────────
    @app.route('/api/thickness/template/<int:equipment_id>', methods=['GET'])
    def download_thickness_template(equipment_id):
        """Descarga una plantilla .xlsx pre-rellenada con los puntos catalogados
        del equipo. El supervisor solo escribe Fecha + Inspector + el valor
        medido (mm) por punto y vuelve a cargar."""
        try:
            from io import BytesIO
            from flask import send_file
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Protection
            from openpyxl.utils import get_column_letter

            eq = Equipment.query.get(equipment_id)
            if not eq:
                return jsonify({"error": f"Equipo {equipment_id} no existe"}), 404

            points = ThicknessPoint.query.filter_by(
                equipment_id=equipment_id, is_active=True
            ).order_by(
                ThicknessPoint.group_name, ThicknessPoint.section,
                ThicknessPoint.order_index, ThicknessPoint.position
            ).all()
            if not points:
                return jsonify({"error": f"El equipo {eq.tag or eq.name} no tiene puntos UT catalogados"}), 400

            wb = Workbook()

            # ── Hoja 1: Datos generales ─────────────────────────────────
            ws1 = wb.active
            ws1.title = "Datos"
            bold = Font(bold=True)
            yellow = PatternFill('solid', fgColor='FFF2CC')  # celdas a llenar
            grey = PatternFill('solid', fgColor='D9D9D9')   # celdas bloqueadas
            ws1['A1'] = 'INSPECCION DE ESPESORES POR ULTRASONIDO'
            ws1['A1'].font = Font(bold=True, size=14)
            ws1.merge_cells('A1:B1')

            ws1['A3'] = 'Equipo:'; ws1['A3'].font = bold
            ws1['B3'] = f"{eq.tag or '-'} — {eq.name}"
            ws1['B3'].fill = grey
            ws1['A4'] = 'equipment_id:'; ws1['A4'].font = bold
            ws1['B4'] = equipment_id
            ws1['B4'].fill = grey
            ws1['A5'] = 'Fecha medicion (YYYY-MM-DD):'; ws1['A5'].font = bold
            ws1['B5'] = dt.date.today().isoformat()
            ws1['B5'].fill = yellow
            ws1['A6'] = 'Inspector:'; ws1['A6'].font = bold
            ws1['B6'] = ''
            ws1['B6'].fill = yellow
            ws1['A7'] = 'Frecuencia (dias):'; ws1['A7'].font = bold
            ws1['B7'] = 60
            ws1['B7'].fill = yellow
            ws1['A8'] = 'Observaciones:'; ws1['A8'].font = bold
            ws1['B8'] = ''
            ws1['B8'].fill = yellow

            ws1['A10'] = 'Instrucciones:'; ws1['A10'].font = bold
            inst = [
                '1. NO modifiques las celdas grises (Equipo, equipment_id, point_id).',
                '2. En las celdas amarillas escribe Fecha, Inspector, Frec. y Obs. (opcional).',
                '3. Ve a la hoja "Mediciones" y escribe el valor medido en mm en la columna G.',
                '4. Si un punto no se midio, deja el valor vacio (se omite, no afecta).',
                '5. Guarda el archivo y subelo desde el modulo de Espesores.',
            ]
            for i, t in enumerate(inst, start=11):
                ws1[f'A{i}'] = t
            ws1.column_dimensions['A'].width = 32
            ws1.column_dimensions['B'].width = 60

            # ── Hoja 2: Mediciones ──────────────────────────────────────
            ws2 = wb.create_sheet('Mediciones')
            headers = ['Grupo', 'Seccion', 'Posicion', 'Nominal (mm)',
                       'Alarma (mm)', 'Scrap (mm)', 'Valor medido (mm)', 'point_id']
            for col, h in enumerate(headers, start=1):
                c = ws2.cell(row=1, column=col, value=h)
                c.font = bold
                c.fill = PatternFill('solid', fgColor='305496')
                c.font = Font(bold=True, color='FFFFFF')
                c.alignment = Alignment(horizontal='center')
            for i, p in enumerate(points, start=2):
                ws2.cell(row=i, column=1, value=p.group_name).fill = grey
                ws2.cell(row=i, column=2, value=p.section).fill = grey
                ws2.cell(row=i, column=3, value=p.position).fill = grey
                ws2.cell(row=i, column=4, value=p.nominal_thickness).fill = grey
                ws2.cell(row=i, column=5, value=p.alarm_thickness).fill = grey
                ws2.cell(row=i, column=6, value=p.scrap_thickness).fill = grey
                # Columna G: valor medido (amarilla, editable)
                cv = ws2.cell(row=i, column=7, value=None)
                cv.fill = yellow
                cv.alignment = Alignment(horizontal='center')
                # Columna H: point_id (gris, NO tocar)
                ws2.cell(row=i, column=8, value=p.id).fill = grey

            for col_idx, w in enumerate([18, 10, 14, 14, 14, 14, 18, 12], start=1):
                ws2.column_dimensions[get_column_letter(col_idx)].width = w
            ws2.freeze_panes = 'A2'

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)

            fname = f"plantilla_UT_{eq.tag or equipment_id}_{dt.date.today().isoformat()}.xlsx"
            return send_file(
                buf,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=fname,
            )
        except Exception as e:
            logger.exception('thickness template download error')
            return jsonify({"error": str(e)}), 500

    @app.route('/api/thickness/upload-template', methods=['POST'])
    def upload_thickness_template():
        """Recibe la plantilla .xlsx rellenada y crea ThicknessInspection +
        ThicknessReadings. Reusa la logica de POST /api/thickness/inspections."""
        try:
            from openpyxl import load_workbook
            if 'file' not in request.files:
                return jsonify({"error": "Falta el archivo (campo 'file')"}), 400
            f = request.files['file']
            if not f.filename.lower().endswith('.xlsx'):
                return jsonify({"error": "El archivo debe ser .xlsx"}), 400

            wb = load_workbook(f, data_only=True)
            if 'Datos' not in wb.sheetnames or 'Mediciones' not in wb.sheetnames:
                return jsonify({"error": "Plantilla invalida: faltan hojas 'Datos' o 'Mediciones'"}), 400

            ws1 = wb['Datos']
            try:
                equipment_id = int(ws1['B4'].value)
            except Exception:
                return jsonify({"error": "Celda B4 (equipment_id) invalida en hoja Datos"}), 400
            eq = Equipment.query.get(equipment_id)
            if not eq:
                return jsonify({"error": f"Equipo {equipment_id} no existe"}), 404
            inspection_date = str(ws1['B5'].value or '').strip()
            if not inspection_date:
                return jsonify({"error": "Falta Fecha medicion (celda B5)"}), 400
            # Excel devuelve datetime si se escribio como fecha
            if hasattr(ws1['B5'].value, 'isoformat'):
                inspection_date = ws1['B5'].value.date().isoformat() if hasattr(ws1['B5'].value, 'date') else ws1['B5'].value.isoformat()
            else:
                # Validar formato YYYY-MM-DD
                try:
                    dt.date.fromisoformat(inspection_date[:10])
                    inspection_date = inspection_date[:10]
                except Exception:
                    return jsonify({"error": f"Fecha invalida '{inspection_date}'. Usa YYYY-MM-DD"}), 400
            inspector = str(ws1['B6'].value or '').strip() or None
            try:
                frequency_days = int(ws1['B7'].value or 60)
            except Exception:
                frequency_days = 60
            observations = str(ws1['B8'].value or '').strip() or None

            # Leer mediciones (saltar fila 1 = headers)
            ws2 = wb['Mediciones']
            readings = []
            row = 2
            while True:
                point_id_cell = ws2.cell(row=row, column=8).value
                value_cell = ws2.cell(row=row, column=7).value
                if point_id_cell is None and value_cell is None:
                    # fila vacia → fin
                    if row > 2 and ws2.cell(row=row + 1, column=8).value is None:
                        break
                    row += 1
                    if row > 5000:
                        break
                    continue
                try:
                    pid = int(point_id_cell)
                except Exception:
                    row += 1
                    continue
                if value_cell is None or str(value_cell).strip() == '':
                    row += 1
                    continue
                try:
                    val = float(value_cell)
                except Exception:
                    row += 1
                    continue
                readings.append({"point_id": pid, "value_mm": val})
                row += 1

            if not readings:
                return jsonify({"error": "No se detecto ninguna medicion en la columna G"}), 400

            # Reusar la logica del POST inspections via llamada interna a la BD
            try:
                insp_dt = dt.date.fromisoformat(inspection_date)
            except Exception:
                insp_dt = dt.date.today()
            next_due = (insp_dt + dt.timedelta(days=frequency_days)).isoformat()

            inspection = ThicknessInspection(
                equipment_id=equipment_id,
                inspection_date=inspection_date,
                next_due_date=next_due,
                frequency_days=frequency_days,
                inspector_name=inspector,
                status='COMPLETA',
                observations=observations,
            )
            db.session.add(inspection)
            db.session.flush()

            total = criticals = alerts = 0
            critical_details = []
            for r in readings:
                pt = ThicknessPoint.query.get(r['point_id'])
                if not pt or pt.equipment_id != equipment_id:
                    continue
                status, is_alert, is_critical = _calc_status(r['value_mm'], pt)
                rd = ThicknessReading(
                    inspection_id=inspection.id,
                    point_id=pt.id,
                    value_mm=r['value_mm'],
                    is_alert=is_alert,
                    is_critical=is_critical,
                )
                db.session.add(rd)
                pt.last_value = r['value_mm']
                pt.last_date = inspection_date
                pt.status = status
                total += 1
                if is_critical:
                    criticals += 1
                    critical_details.append(
                        f"{pt.group_name} S{pt.section or ''}-{pt.position}: "
                        f"{r['value_mm']}mm (limite {pt.scrap_thickness}mm)"
                    )
                elif is_alert:
                    alerts += 1

            inspection.total_points = total
            inspection.critical_points = criticals
            inspection.alert_points = alerts
            if criticals > 0:
                inspection.semaphore_status = 'ROJO'
            elif alerts > 0:
                inspection.semaphore_status = 'AMARILLO'
            else:
                inspection.semaphore_status = 'VERDE'
            db.session.commit()

            # Aviso automatico si hay criticos (mismo patron que el POST inspections)
            notice_code = None
            if criticals > 0 and MaintenanceNotice:
                try:
                    desc = (f"Inspeccion UT detecto {criticals} punto(s) critico(s) en "
                            f"{eq.name}.\nDetalles:\n- " + "\n- ".join(critical_details[:10]))
                    notice = MaintenanceNotice(
                        description=desc, equipment_id=equipment_id,
                        criticality='Alta', priority='Alta',
                        maintenance_type='Correctivo', failure_category='Estructural',
                        failure_mode='Desgaste', status='Pendiente', scope='PLAN',
                        reporter_name=inspector or 'Plantilla UT',
                    )
                    db.session.add(notice)
                    db.session.flush()
                    notice.code = f"AV-{notice.id:04d}"
                    notice_code = notice.code
                    db.session.commit()
                except Exception as ne:
                    logger.warning(f"thickness upload: aviso auto fallo: {ne}")
                    db.session.rollback()

            return jsonify({
                "ok": True,
                "inspection_id": inspection.id,
                "equipment_tag": eq.tag,
                "equipment_name": eq.name,
                "inspection_date": inspection_date,
                "total_readings": total,
                "criticals": criticals,
                "alerts": alerts,
                "semaphore_status": inspection.semaphore_status,
                "notice_code": notice_code,
            }), 201
        except Exception as e:
            db.session.rollback()
            logger.exception('thickness upload template error')
            return jsonify({"error": str(e)}), 500
