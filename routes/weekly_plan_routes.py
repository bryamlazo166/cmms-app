"""Programa Nocturno Semanal — preventivos ejecutados por proveedor turno noche.

Flujo:
  1. Jefe de mtto crea un plan para una semana (lunes-domingo).
  2. Auto-planner distribuye puntos preventivos vencidos/próximos en las 4
     áreas × 7 noches, llenando la capacidad horaria (default 2 técnicos × 12 h).
  3. Plan se publica → genera token URL-safe para el proveedor.
  4. Proveedor abre /programa-nocturno/publico/<token> desde su móvil,
     marca ítems como ejecutados con observaciones.
  5. Cada ítem ejecutado crea OT con source_type/source_id → el handler
     _update_source_on_close actualiza last_service_date y next_due_date.

Estimaciones de duración por tipo de tarea (si no vienen explícitas):
  - Lubricación: 0.5 h base + 0.1 h por litro nominal
  - Inspección: 1.0 h base + 0.1 h por ítem activo
  - Monitoreo: 0.5 h base
"""
import datetime as dt
import secrets
from io import BytesIO

from flask import jsonify, request, render_template, send_file

from utils.specialty_helpers import discipline_for_weekly_item


# ── Estimación de duración por tarea ─────────────────────────────────────────

def _estimate_hours(source_type, obj, WorkOrder=None):
    """Estima horas de una tarea preventiva.

    Estrategia:
      1. Si hay ≥2 OTs cerradas del mismo source → promedio de las últimas 3
         (rolling average auto-aprendiente)
      2. Si no hay suficiente histórico → heurística por tipo.
    """
    source_id = getattr(obj, 'id', None)
    if WorkOrder is not None and source_type and source_id:
        try:
            recent = WorkOrder.query.filter(
                WorkOrder.source_type == source_type,
                WorkOrder.source_id == source_id,
                WorkOrder.status == 'Cerrada',
                WorkOrder.real_duration.isnot(None),
                WorkOrder.real_duration > 0,
            ).order_by(WorkOrder.id.desc()).limit(3).all()
            if len(recent) >= 2:
                avg = sum(float(w.real_duration) for w in recent) / len(recent)
                # Caps defensivos: no menor a 0.25h ni mayor a 8h
                return round(max(0.25, min(avg, 8.0)), 2)
        except Exception:
            pass

    # Fallback heurístico (primer preventivo sin historial)
    if source_type == 'lubrication':
        base = 0.5
        qty = getattr(obj, 'quantity_nominal', None) or 0
        return round(base + min(qty * 0.1, 2.0), 2)
    if source_type == 'inspection':
        base = 1.0
        items_count = 0
        if hasattr(obj, 'items') and obj.items:
            items_count = len([i for i in obj.items if i.is_active])
        return round(base + min(items_count * 0.08, 1.5), 2)
    if source_type == 'monitoring':
        return 0.5
    return 1.0


# ── Código e ISO week helpers ────────────────────────────────────────────────

def _iso_week_bounds(date_str):
    """Devuelve (lunes, domingo) de la semana que contiene a date_str."""
    d = dt.date.fromisoformat(date_str)
    monday = d - dt.timedelta(days=d.weekday())
    sunday = monday + dt.timedelta(days=6)
    return monday.isoformat(), sunday.isoformat()


def _generate_plan_code(week_start, WeeklyPlan):
    """PN-YYYY-WW (ISO week)."""
    try:
        d = dt.date.fromisoformat(week_start)
        year, week, _ = d.isocalendar()
        return f"PN-{year}-{week:02d}"
    except Exception:
        return f"PN-{dt.date.today().strftime('%Y-%W')}"


def register_weekly_plan_routes(
    app, db, logger,
    WeeklyPlan, WeeklyPlanItem,
    Area, Line, Equipment, WorkOrder, Provider,
    LubricationPoint, InspectionRoute, MonitoringPoint,
    _calculate_lubrication_schedule, _calculate_monitoring_schedule,
):

    # ── Página web (solo dashboard del jefe) ─────────────────────────────

    @app.route('/programa-nocturno', methods=['GET'])
    def wp_page():
        return render_template('programa_nocturno.html')

    @app.route('/api/preventive-sources', methods=['GET'])
    def wp_list_preventive_sources():
        """Endpoint auxiliar para el selector de ítems manuales.
        Filtros: ?source_type=lubrication|inspection|monitoring&area_id=N
        """
        try:
            from utils.preventive_sources import collect_sources
            src_type = request.args.get('source_type')
            area_id = request.args.get('area_id', type=int)
            sources = collect_sources(
                LubricationPoint, InspectionRoute, MonitoringPoint,
                _calc_lub_schedule=_calculate_lubrication_schedule,
                _calc_mon_schedule=_calculate_monitoring_schedule,
                source_types={src_type} if src_type else None,
                area_ids=[area_id] if area_id else None,
                enrich_names=True,
            )
            return jsonify(sources)
        except Exception as e:
            logger.error(f"list_preventive_sources error: {e}")
            return jsonify({"error": str(e)}), 500

    # ── CRUD de planes ───────────────────────────────────────────────────

    @app.route('/api/weekly-plans', methods=['GET'])
    def wp_list():
        try:
            plans = WeeklyPlan.query.order_by(WeeklyPlan.week_start.desc()).limit(20).all()
            out = []
            for p in plans:
                d = p.to_dict()
                d['item_count'] = len(p.items)
                d['executed_count'] = sum(1 for it in p.items if it.status == 'EJECUTADO')
                d['total_hours'] = round(sum(it.estimated_hours for it in p.items), 2)
                out.append(d)
            return jsonify(out)
        except Exception as e:
            logger.error(f"list_weekly_plans error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans', methods=['POST'])
    def wp_create():
        try:
            data = request.get_json() or {}
            ref_date = data.get('week_start') or dt.date.today().isoformat()
            mon, sun = _iso_week_bounds(ref_date)

            # ¿Ya existe plan para esa semana?
            existing = WeeklyPlan.query.filter_by(week_start=mon).first()
            if existing:
                return jsonify({"error": f"Ya existe un plan para la semana del {mon}", "id": existing.id}), 400

            plan = WeeklyPlan(
                week_start=mon,
                week_end=sun,
                provider_id=data.get('provider_id'),
                tech_count=int(data.get('tech_count', 2)),
                hours_per_night=float(data.get('hours_per_night', 12.0)),
                notes=data.get('notes'),
                created_by=data.get('created_by'),
            )
            db.session.add(plan)
            db.session.flush()
            plan.code = _generate_plan_code(mon, WeeklyPlan)
            db.session.commit()
            return jsonify(plan.to_dict(include_items=True)), 201
        except Exception as e:
            db.session.rollback()
            logger.error(f"create_weekly_plan error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans/<int:plan_id>', methods=['GET'])
    def wp_detail(plan_id):
        try:
            p = WeeklyPlan.query.get_or_404(plan_id)
            d = p.to_dict(include_items=True)
            # Agrupar ítems por día × área para vista calendario
            grid = {}
            for it in p.items:
                key = (it.day_of_week, it.area_id)
                grid.setdefault(str(key), []).append(it.to_dict())
            # Horas por día (suma por día_of_week)
            hours_per_day = [0.0] * 7
            for it in p.items:
                if 0 <= it.day_of_week <= 6:
                    hours_per_day[it.day_of_week] += it.estimated_hours or 0
            d['grid'] = grid
            d['hours_per_day'] = [round(h, 2) for h in hours_per_day]
            d['capacity_per_day'] = round(p.tech_count * p.hours_per_night, 2)
            d['executed_count'] = sum(1 for it in p.items if it.status == 'EJECUTADO')
            d['total_hours'] = round(sum(it.estimated_hours for it in p.items), 2)
            return jsonify(d)
        except Exception as e:
            logger.error(f"get_weekly_plan error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans/<int:plan_id>', methods=['PUT'])
    def wp_update(plan_id):
        try:
            p = WeeklyPlan.query.get_or_404(plan_id)
            data = request.get_json() or {}
            for field in ('notes', 'status', 'created_by'):
                if field in data:
                    setattr(p, field, data[field])
            if 'provider_id' in data:
                p.provider_id = int(data['provider_id']) if data['provider_id'] else None
            if 'tech_count' in data:
                p.tech_count = int(data['tech_count'])
            if 'hours_per_night' in data:
                p.hours_per_night = float(data['hours_per_night'])
            db.session.commit()
            return jsonify(p.to_dict())
        except Exception as e:
            db.session.rollback()
            logger.error(f"update_weekly_plan error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans/<int:plan_id>', methods=['DELETE'])
    def wp_delete(plan_id):
        try:
            p = WeeklyPlan.query.get_or_404(plan_id)
            db.session.delete(p)
            db.session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            db.session.rollback()
            logger.error(f"delete_weekly_plan error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans/<int:plan_id>/publish', methods=['POST'])
    def wp_publish(plan_id):
        """Cambia estado a PUBLICADO y genera token público para el proveedor."""
        try:
            p = WeeklyPlan.query.get_or_404(plan_id)
            if not p.public_token:
                p.public_token = secrets.token_urlsafe(32)
            p.status = 'PUBLICADO'
            db.session.commit()
            host = request.host_url.rstrip('/')
            public_url = f"{host}/programa-nocturno/publico/{p.public_token}"
            return jsonify({
                "ok": True,
                "public_token": p.public_token,
                "public_url": public_url,
                "plan": p.to_dict(),
            })
        except Exception as e:
            db.session.rollback()
            logger.error(f"publish_weekly_plan error: {e}")
            return jsonify({"error": str(e)}), 500

    # ── Auto-planner ─────────────────────────────────────────────────────

    @app.route('/api/weekly-plans/<int:plan_id>/auto-plan', methods=['POST'])
    def wp_auto_plan(plan_id):
        """Distribuye puntos preventivos en las 4 áreas × 7 noches llenando la capacidad.

        Algoritmo:
        - Recolecta puntos preventivos activos.
        - FILTRO POR RESPONSABLE: si el plan tiene provider_id, solo incluye puntos
          cuya responsabilidad efectiva sea ese proveedor (matriz de
          responsabilidad: Equipment.default_responsible_party + override por punto).
        - GRUPO POR EQUIPO: agrupa los puntos por (area, equipment) para que TODOS
          los puntos del mismo equipo caigan el MISMO dia (preferencia del usuario:
          terminar un equipo antes de pasar al siguiente — el lubricador ataca el
          equipo completo en una visita).
        - Cada grupo se asigna al dia con menos carga dentro de su area.

        Body opcional: { clear_existing: bool, only_provider: bool, group_by_equipment: bool }
        """
        try:
            from utils.preventive_sources import collect_sources
            from utils.responsibility import resolve_responsibility, INTERNO, PROVEEDOR

            p = WeeklyPlan.query.get_or_404(plan_id)
            data = request.get_json() or {}
            clear_existing = bool(data.get('clear_existing', True))
            # Por defecto si el plan tiene proveedor, filtramos por el (es lo
            # que el usuario espera: el plan del proveedor solo trae lo del proveedor).
            only_provider = bool(data.get('only_provider', bool(p.provider_id)))
            group_by_equipment = bool(data.get('group_by_equipment', True))

            if clear_existing:
                for it in list(p.items):
                    if it.status == 'PLANIFICADO':
                        db.session.delete(it)
                db.session.flush()

            capacity_per_day = p.tech_count * p.hours_per_night

            areas_all = Area.query.all()
            line_map = {l.id: l for l in Line.query.all()}
            equip_map = {e.id: e for e in Equipment.query.all()}
            area_map = {a.id: a for a in areas_all}

            open_ots = WorkOrder.query.filter(
                WorkOrder.status.in_(['Abierta', 'Programada', 'En Progreso']),
                WorkOrder.source_type.isnot(None),
            ).all()
            excluded = {(o.source_type, o.source_id) for o in open_ots if o.source_id}
            for it in p.items:
                if it.source_type and it.source_id:
                    excluded.add((it.source_type, it.source_id))

            sources = collect_sources(
                LubricationPoint, InspectionRoute, MonitoringPoint,
                _calc_lub_schedule=_calculate_lubrication_schedule,
                _calc_mon_schedule=_calculate_monitoring_schedule,
                only_overdue=False,
                exclude=excluded,
                enrich_names=True,
                area_map=area_map, line_map=line_map, equip_map=equip_map,
            )

            # ── Filtrado por responsabilidad ────────────────────────────────
            # El plan apunta a UN proveedor (o ninguno). Filtramos los puntos
            # cuya responsabilidad efectiva sea ese proveedor (o cualquiera si
            # only_provider=False, util si quieres ver carga total).
            def _get_source_obj(src):
                t, sid = src['source_type'], src['source_id']
                if t == 'lubrication':
                    return LubricationPoint.query.get(sid)
                if t == 'inspection':
                    return InspectionRoute.query.get(sid)
                if t == 'monitoring':
                    return MonitoringPoint.query.get(sid)
                return None

            # Enriquecer todos los sources con responsabilidad
            for src in sources:
                obj = _get_source_obj(src)
                if not obj:
                    src['_obj'] = None
                    continue
                eq = equip_map.get(getattr(obj, 'equipment_id', None))
                party, prov_id = resolve_responsibility(obj, equipment=eq)
                src['_obj'] = obj
                src['_party'] = party
                src['_provider_id'] = prov_id
                src['_equipment_id'] = getattr(obj, 'equipment_id', None)
            sources = [s for s in sources if s.get('_obj')]

            # ── Filtrado por responsabilidad (con fallback inteligente) ─────
            # Si el plan tiene proveedor y only_provider=True, intentamos
            # filtrar. Pero si el resultado deja CERO puntos (porque el
            # usuario aun no asigno equipos al proveedor), hacemos FALLBACK
            # automatico: usamos todos los puntos y avisamos en la respuesta.
            warning_msg = None
            total_before_filter = len(sources)
            if only_provider and p.provider_id:
                provider_sources = []
                for src in sources:
                    party = src.get('_party')
                    prov_id = src.get('_provider_id')
                    if party != PROVEEDOR:
                        continue
                    if prov_id and prov_id != p.provider_id:
                        continue
                    provider_sources.append(src)
                if provider_sources:
                    sources = provider_sources
                else:
                    # Fallback: nadie esta marcado como del proveedor todavia.
                    warning_msg = (
                        f"No hay puntos preventivos asignados al proveedor del plan. "
                        f"Se planificaron TODOS los {total_before_filter} puntos disponibles. "
                        f"Para que el filtro funcione, asigna primero los equipos al "
                        f"proveedor en /responsabilidades (o edita Equipment.default_responsible_party)."
                    )
                    only_provider = False  # marcar como no-filtrado para la respuesta

            # ── Agrupar por equipo si group_by_equipment ────────────────────
            # Cada grupo se considera una unidad atomica que cae el mismo dia
            # en la misma area. El orden de procesamiento es por prioridad del
            # peor punto del grupo (semaforo mas critico).
            if group_by_equipment:
                from collections import OrderedDict
                groups_dict = OrderedDict()
                # Mantener orden de llegada para preservar prioridad de collect_sources
                for src in sources:
                    eq_id = src.get('_equipment_id') or 0
                    aid = src.get('area_id') or 0
                    key = (aid, eq_id)
                    groups_dict.setdefault(key, []).append(src)
                groups = list(groups_dict.values())
            else:
                groups = [[s] for s in sources]

            # Carga inicial por dia/area (considerando items EJECUTADO ya en plan)
            load = {d: {a.id: 0.0 for a in areas_all} for d in range(7)}
            day_totals = [0.0] * 7
            for it in p.items:
                if 0 <= it.day_of_week <= 6 and it.area_id in load[it.day_of_week]:
                    load[it.day_of_week][it.area_id] += it.estimated_hours or 0
                    day_totals[it.day_of_week] += it.estimated_hours or 0

            placed = 0
            skipped = 0
            for group in groups:
                if not group or not group[0].get('area_id'):
                    continue
                aid = group[0]['area_id']

                # Calcular horas totales del grupo
                group_hours = []
                for src in group:
                    obj = src.get('_obj')
                    h = _estimate_hours(src['source_type'], obj, WorkOrder=WorkOrder) if obj else 1.0
                    group_hours.append(h)
                total_group_hours = sum(group_hours)

                # Elegir dia con menos carga en el area, y donde quepa el grupo COMPLETO
                candidate_days = list(range(7))
                candidate_days.sort(key=lambda d: (load[d].get(aid, 0), day_totals[d]))

                chosen_day = None
                for d in candidate_days:
                    if day_totals[d] + total_group_hours <= capacity_per_day:
                        chosen_day = d
                        break
                if chosen_day is None:
                    # No cabe el grupo completo en ningun dia.
                    # Politica: si group_by_equipment, NO partir el grupo (preferencia
                    # del usuario: mejor postergar todo el equipo). Skip.
                    skipped += len(group)
                    if min(day_totals) >= capacity_per_day:
                        break
                    continue

                # Colocar todos los items del grupo en chosen_day
                order_base = sum(1 for it in p.items if it.day_of_week == chosen_day and it.area_id == aid)
                for i, src in enumerate(group):
                    item = WeeklyPlanItem(
                        plan_id=p.id,
                        day_of_week=chosen_day,
                        area_id=aid,
                        order_index=order_base + i,
                        source_type=src['source_type'],
                        source_id=src['source_id'],
                        source_code=src.get('code', ''),
                        source_name=src.get('name', ''),
                        equipment_tag=src.get('equipment_tag', ''),
                        description=src.get('description', ''),
                        estimated_hours=group_hours[i],
                        status='PLANIFICADO',
                    )
                    db.session.add(item)
                    p.items.append(item)
                    placed += 1
                load[chosen_day][aid] = load[chosen_day].get(aid, 0) + total_group_hours
                day_totals[chosen_day] += total_group_hours

                if all(dt_h >= capacity_per_day for dt_h in day_totals):
                    break

            db.session.commit()
            return jsonify({
                "ok": True,
                "items_placed": placed,
                "items_skipped_no_capacity": skipped,
                "groups_processed": len(groups),
                "filtered_by_provider": bool(only_provider and p.provider_id),
                "grouped_by_equipment": group_by_equipment,
                "hours_per_day": [round(x, 2) for x in day_totals],
                "capacity_per_day": capacity_per_day,
                "fill_pct": [round((x / capacity_per_day * 100) if capacity_per_day else 0, 1) for x in day_totals],
                "warning": warning_msg,
            })
        except Exception as e:
            db.session.rollback()
            logger.error(f"auto_plan_week error: {e}")
            import traceback; traceback.print_exc()
            return jsonify({"error": str(e)}), 500

    # ── CRUD de ítems ────────────────────────────────────────────────────

    @app.route('/api/weekly-plans/<int:plan_id>/items', methods=['POST'])
    def wp_item_add(plan_id):
        """Agregar ítem manual al plan (sin pasar por auto-planner)."""
        try:
            p = WeeklyPlan.query.get_or_404(plan_id)
            data = request.get_json() or {}
            item = WeeklyPlanItem(
                plan_id=p.id,
                day_of_week=int(data.get('day_of_week', 0)),
                area_id=data.get('area_id'),
                source_type=data.get('source_type'),
                source_id=data.get('source_id'),
                source_code=data.get('source_code'),
                source_name=data.get('source_name'),
                equipment_tag=data.get('equipment_tag'),
                description=data.get('description') or '',
                estimated_hours=float(data.get('estimated_hours', 1.0)),
                order_index=int(data.get('order_index', 99)),
            )
            db.session.add(item)
            db.session.commit()
            return jsonify(item.to_dict()), 201
        except Exception as e:
            db.session.rollback()
            logger.error(f"add_plan_item error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans/<int:plan_id>/items/<int:item_id>', methods=['PUT'])
    def wp_item_update(plan_id, item_id):
        """Actualizar ítem (usado para drag & drop y edición manual)."""
        try:
            item = WeeklyPlanItem.query.filter_by(id=item_id, plan_id=plan_id).first_or_404()
            data = request.get_json() or {}
            if 'day_of_week' in data:
                item.day_of_week = int(data['day_of_week'])
            if 'area_id' in data:
                item.area_id = int(data['area_id']) if data['area_id'] else None
            if 'order_index' in data:
                item.order_index = int(data['order_index'])
            if 'estimated_hours' in data:
                item.estimated_hours = float(data['estimated_hours'])
            if 'description' in data:
                item.description = data['description']
            db.session.commit()
            return jsonify(item.to_dict())
        except Exception as e:
            db.session.rollback()
            logger.error(f"update_plan_item error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans/<int:plan_id>/items/<int:item_id>', methods=['DELETE'])
    def wp_item_delete(plan_id, item_id):
        try:
            item = WeeklyPlanItem.query.filter_by(id=item_id, plan_id=plan_id).first_or_404()
            db.session.delete(item)
            db.session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            db.session.rollback()
            logger.error(f"delete_plan_item error: {e}")
            return jsonify({"error": str(e)}), 500

    # ── Ejecución: crea OT vinculada + actualiza source ──────────────────

    def _execute_item(item, plan, executed_by=None, notes=None):
        """Marca ítem como EJECUTADO y crea OT cerrada vinculada al source."""
        from models import MaintenanceNotice
        today = dt.datetime.now().strftime('%Y-%m-%d')

        # Resolver jerarquía
        area_id = item.area_id
        line_id = None
        equipment_id = None
        system_id = None
        component_id = None
        if item.source_type == 'lubrication' and item.source_id:
            pt = LubricationPoint.query.get(item.source_id)
            if pt:
                area_id = area_id or pt.area_id
                line_id = pt.line_id
                equipment_id = pt.equipment_id
                system_id = pt.system_id
                component_id = pt.component_id
        elif item.source_type == 'inspection' and item.source_id:
            rt = InspectionRoute.query.get(item.source_id)
            if rt:
                area_id = area_id or rt.area_id
                line_id = rt.line_id
                equipment_id = rt.equipment_id
        elif item.source_type == 'monitoring' and item.source_id:
            pt = MonitoringPoint.query.get(item.source_id)
            if pt:
                area_id = area_id or pt.area_id
                line_id = pt.line_id
                equipment_id = pt.equipment_id
                system_id = pt.system_id
                component_id = pt.component_id

        # Crear OT cerrada directamente (preventivo ejecutado por proveedor)
        wo = WorkOrder(
            description=item.description or item.source_name or '(preventivo)',
            maintenance_type='Preventivo',
            status='Cerrada',
            real_start_date=today,
            real_end_date=today,
            real_duration=item.estimated_hours,
            estimated_duration=item.estimated_hours,
            tech_count=plan.tech_count,
            area_id=area_id,
            line_id=line_id,
            equipment_id=equipment_id,
            system_id=system_id,
            component_id=component_id,
            source_type=item.source_type,
            source_id=item.source_id,
            execution_comments=notes or '',
            provider_id=plan.provider_id,
        )
        db.session.add(wo)
        db.session.flush()
        wo.code = f"OT-{wo.id:04d}"

        # Actualizar source (reutiliza misma lógica que work_orders_routes)
        try:
            if item.source_type == 'lubrication' and item.source_id:
                pt = LubricationPoint.query.get(item.source_id)
                if pt:
                    pt.last_service_date = today
                    if _calculate_lubrication_schedule:
                        nd, sem = _calculate_lubrication_schedule(today, pt.frequency_days, pt.warning_days)
                        pt.next_due_date = nd
                        pt.semaphore_status = sem
            elif item.source_type == 'inspection' and item.source_id:
                rt = InspectionRoute.query.get(item.source_id)
                if rt:
                    rt.last_execution_date = today
                    if _calculate_lubrication_schedule:
                        nd, sem = _calculate_lubrication_schedule(today, rt.frequency_days, rt.warning_days)
                        rt.next_due_date = nd
                        rt.semaphore_status = sem
            elif item.source_type == 'monitoring' and item.source_id:
                pt = MonitoringPoint.query.get(item.source_id)
                if pt:
                    pt.last_measurement_date = today
                    if _calculate_monitoring_schedule:
                        nd, sem = _calculate_monitoring_schedule(today, pt.frequency_days, pt.warning_days)
                        pt.next_due_date = nd
                        pt.semaphore_status = sem
        except Exception as upd_err:
            logger.warning(f"No se pudo actualizar source: {upd_err}")

        # Marcar ítem
        item.status = 'EJECUTADO'
        item.executed_at = today
        item.executed_by = executed_by or 'proveedor'
        item.execution_notes = notes or ''
        item.work_order_id = wo.id
        return wo

    @app.route('/api/weekly-plans/<int:plan_id>/items/<int:item_id>/execute', methods=['POST'])
    def wp_item_execute(plan_id, item_id):
        """Marcar ítem como ejecutado → crea OT y actualiza el punto origen."""
        try:
            item = WeeklyPlanItem.query.filter_by(id=item_id, plan_id=plan_id).first_or_404()
            plan = WeeklyPlan.query.get(plan_id)
            data = request.get_json() or {}
            if item.status == 'EJECUTADO':
                return jsonify({"error": "El ítem ya fue ejecutado", "work_order_id": item.work_order_id}), 400
            wo = _execute_item(item, plan,
                               executed_by=data.get('executed_by'),
                               notes=data.get('notes'))
            db.session.commit()
            return jsonify({
                "ok": True,
                "item": item.to_dict(),
                "work_order_code": wo.code,
            })
        except Exception as e:
            db.session.rollback()
            logger.error(f"execute_plan_item error: {e}")
            return jsonify({"error": str(e)}), 500

    # ── Vista pública tokenizada (para el proveedor, sin login) ──────────

    @app.route('/programa-nocturno/publico/<token>', methods=['GET'])
    def wp_public_view(token):
        p = WeeklyPlan.query.filter_by(public_token=token).first()
        if not p:
            return "Enlace inválido o plan cerrado.", 404
        return render_template('programa_nocturno_publico.html', token=token)

    @app.route('/api/public/weekly-plans/<token>', methods=['GET'])
    def wp_public_data(token):
        """API pública: datos del plan por token (sin login)."""
        try:
            p = WeeklyPlan.query.filter_by(public_token=token).first()
            if not p:
                return jsonify({"error": "Token inválido"}), 404
            d = p.to_dict(include_items=True)
            grid = {}
            for it in p.items:
                key = str((it.day_of_week, it.area_id))
                grid.setdefault(key, []).append(it.to_dict())
            d['grid'] = grid
            return jsonify(d)
        except Exception as e:
            logger.error(f"public_weekly_plan_data error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/public/weekly-plans/<token>/items/<int:item_id>/execute', methods=['POST'])
    def wp_public_execute(token, item_id):
        """Proveedor marca ítem ejecutado desde el link público."""
        try:
            p = WeeklyPlan.query.filter_by(public_token=token).first()
            if not p:
                return jsonify({"error": "Token inválido"}), 404
            item = WeeklyPlanItem.query.filter_by(id=item_id, plan_id=p.id).first_or_404()
            if item.status == 'EJECUTADO':
                return jsonify({"ok": True, "already_executed": True})
            data = request.get_json() or {}
            wo = _execute_item(item, p,
                               executed_by=data.get('executed_by') or 'proveedor',
                               notes=data.get('notes'))
            db.session.commit()
            return jsonify({
                "ok": True,
                "item": item.to_dict(),
                "work_order_code": wo.code,
            })
        except Exception as e:
            db.session.rollback()
            logger.error(f"public_execute_item error: {e}")
            return jsonify({"error": str(e)}), 500

    # ── Vista Matriz: actividades x dias (estilo Excel del proveedor) ─────
    # El usuario maneja un Excel donde cada fila es (Area, Equipo, Sub-Equipo,
    # Actividad) y las columnas son los 7 dias con celdas marcadas con "P"
    # cuando esa actividad debe ejecutarse ese dia. Esta vista replica ese
    # formato dentro del CMMS y permite exportarlo.
    #
    # Consolidacion inteligente: cuando el mismo equipo/dia tiene 3 puntos
    # del tipo "lubricacion chumacera motriz", "lubricacion chumacera
    # conducida" y "lubricacion cadena" → se consolidan en una sola fila
    # "LUBRICACION DE CHUMACERAS Y CADENA" (porque en la realidad es UNA
    # sola visita del lubricador que toca los 3 puntos).

    def _extract_activity_parts(source_name):
        """Extrae (verbo, [componentes]) de un nombre tipo
        'LUBRICACION CHUMACERA MOTRIZ TH1'.
        Retorna ('LUBRICACION', ['CHUMACERA MOTRIZ']) o (None, [name]) si no
        reconoce el verbo."""
        if not source_name:
            return None, []
        s = str(source_name).upper().strip()
        # Verbos comunes en planes preventivos (orden: mas largo primero)
        verbs = ['LUBRICACION DE', 'LUBRICACION', 'LUBRICACIÓN DE', 'LUBRICACIÓN',
                 'INSPECCION DE', 'INSPECCION', 'INSPECCIÓN DE', 'INSPECCIÓN',
                 'CAMBIO DE', 'CAMBIO',
                 'VERIFICACION DE', 'VERIFICACION', 'VERIFICACIÓN DE', 'VERIFICACIÓN',
                 'MEDICION DE', 'MEDICION', 'MEDICIÓN DE', 'MEDICIÓN',
                 'LIMPIEZA DE', 'LIMPIEZA',
                 'AJUSTE DE', 'AJUSTE',
                 'TOMA DE', 'ANALISIS DE', 'ANÁLISIS DE']
        verb_found = None
        rest = s
        for v in verbs:
            if s.startswith(v + ' '):
                verb_found = v.replace(' DE', '').replace('LUBRICACIÓN', 'LUBRICACION')\
                    .replace('INSPECCIÓN', 'INSPECCION').replace('VERIFICACIÓN', 'VERIFICACION')\
                    .replace('MEDICIÓN', 'MEDICION').replace('ANÁLISIS', 'ANALISIS')
                rest = s[len(v):].strip()
                break
        if not verb_found:
            return None, [s]
        # Quitar tag de equipo del final (ej "TH1", "SEC2-TH1") — usualmente el
        # ultimo token alfanumerico que parece codigo. Heuristica simple:
        # tokens que matchean ^[A-Z]+\d+$ o tienen guion al final.
        import re as _re
        tokens = rest.split()
        while tokens and (_re.match(r'^[A-Z]*\d+[A-Z\d-]*$', tokens[-1])
                          or tokens[-1] in ('Y', 'DE', 'DEL', 'LA', 'EL')):
            tokens.pop()
        component = ' '.join(tokens).strip()
        return verb_found, ([component] if component else [])

    def _consolidate_activity_label(items):
        """Dada una lista de items con mismo verbo+equipo+dia, devuelve UN label
        consolidado tipo 'LUBRICACION DE CHUMACERAS Y CADENA'."""
        if not items:
            return ''
        if len(items) == 1:
            return items[0].source_name or items[0].description or ''

        verb = None
        components = []
        for it in items:
            v, comps = _extract_activity_parts(it.source_name or it.description or '')
            if v and not verb:
                verb = v
            components.extend(comps)

        if not verb:
            # No se pudo extraer verbo; devolver join simple
            return ' + '.join(it.source_name or '' for it in items)

        # Pluralizar y deduplicar componentes
        # Ej: ['CHUMACERA MOTRIZ', 'CHUMACERA CONDUCIDA', 'CADENA']
        #     -> agrupar por palabra raiz: 'CHUMACERAS' + 'CADENA'
        from collections import OrderedDict
        roots = OrderedDict()  # raiz -> lista de calificadores
        for c in components:
            if not c:
                continue
            words = c.split()
            root = words[0] if words else c
            qualifier = ' '.join(words[1:]) if len(words) > 1 else ''
            roots.setdefault(root, []).append(qualifier)

        parts = []
        for root, quals in roots.items():
            quals = [q for q in quals if q]
            if len(quals) > 1 or (len(quals) == 1 and any(q == '' for q in roots[root])):
                # Plural si hay multiples calificadores
                plural = root + 'S' if not root.endswith('S') else root
                parts.append(plural)
            elif len(quals) == 1 and quals[0]:
                parts.append(f"{root} {quals[0]}")
            else:
                parts.append(root)

        if len(parts) == 1:
            return f"{verb} DE {parts[0]}"
        if len(parts) == 2:
            return f"{verb} DE {parts[0]} Y {parts[1]}"
        return f"{verb} DE {', '.join(parts[:-1])} Y {parts[-1]}"

    def _build_matrix(plan):
        """Devuelve la matriz tipo Gantt para un plan:
        rows = [{area, line, equipment_tag, equipment_name, activity, hours, days, responsible_party}]
        days_meta = [{idx, name, date}]
        """
        from models import Equipment as _Eq, Area as _Ar, Line as _Ln, LubricationPoint as _LP, \
            InspectionRoute as _IR, MonitoringPoint as _MP
        from utils.responsibility import resolve_responsibility

        # Agrupar items por (area_id, equipment_tag, verbo) para consolidar
        from collections import defaultdict
        groups = defaultdict(lambda: defaultdict(list))  # key -> day -> [items]
        for it in plan.items:
            verb, _comps = _extract_activity_parts(it.source_name or it.description or '')
            verb_key = verb or 'OTRO'
            tag = it.equipment_tag or '(Sin equipo)'
            area_id = it.area_id or 0
            key = (area_id, tag, verb_key)
            groups[key][it.day_of_week].append(it)

        # Pre-cargar mapas
        area_map = {a.id: a for a in _Ar.query.all()}
        line_map = {l.id: l for l in _Ln.query.all()}
        all_tags = {tag for (_, tag, _) in groups.keys() if tag != '(Sin equipo)'}
        eq_by_tag = {e.tag: e for e in _Eq.query.filter(_Eq.tag.in_(all_tags)).all()} if all_tags else {}
        # Pre-cargar puntos preventivos por tipo+id para resolver responsabilidad
        ids_by_type = defaultdict(set)
        for it in plan.items:
            if it.source_type and it.source_id:
                ids_by_type[it.source_type].add(it.source_id)
        points_map = {}  # (type, id) -> obj
        if ids_by_type.get('lubrication'):
            for o in _LP.query.filter(_LP.id.in_(ids_by_type['lubrication'])).all():
                points_map[('lubrication', o.id)] = o
        if ids_by_type.get('inspection'):
            for o in _IR.query.filter(_IR.id.in_(ids_by_type['inspection'])).all():
                points_map[('inspection', o.id)] = o
        if ids_by_type.get('monitoring'):
            for o in _MP.query.filter(_MP.id.in_(ids_by_type['monitoring'])).all():
                points_map[('monitoring', o.id)] = o

        rows = []
        for (area_id, tag, verb_key), days_dict in groups.items():
            area = area_map.get(area_id)
            eq = eq_by_tag.get(tag)
            line = line_map.get(eq.line_id) if eq and eq.line_id else None
            # Para cada dia consolidamos los items de ese grupo
            day_labels = {}
            day_items = {}
            day_status = {}
            day_hours = {}
            total_hours = 0.0
            unified_label_set = set()
            party_set = set()
            providers_set = set()
            for d, items in days_dict.items():
                label = _consolidate_activity_label(items)
                day_labels[d] = label
                day_items[d] = [it.id for it in items]
                day_status[d] = 'EJECUTADO' if all(it.status == 'EJECUTADO' for it in items) \
                    else ('PARCIAL' if any(it.status == 'EJECUTADO' for it in items) else 'PLANIFICADO')
                hh = sum(float(it.estimated_hours or 0) for it in items)
                day_hours[d] = round(hh, 2)
                total_hours += hh
                unified_label_set.add(label)
                # Resolver responsable para cada item
                for it in items:
                    obj = points_map.get((it.source_type, it.source_id))
                    if obj:
                        p_party, p_pid = resolve_responsibility(obj, equipment=eq)
                        party_set.add(p_party)
                        if p_pid:
                            providers_set.add(p_pid)

            activity_label = max(unified_label_set, key=len) if unified_label_set else verb_key

            # Resumen del responsable: si hay UNO solo lo muestra; si hay mezcla, "MIXTO"
            if len(party_set) == 1:
                row_party = next(iter(party_set))
            elif party_set:
                row_party = 'MIXTO'
            else:
                row_party = ''

            rows.append({
                'area_id': area_id,
                'area_name': area.name if area else '(Sin area)',
                'line_id': line.id if line else None,
                'line_name': line.name if line else '-',
                'equipment_tag': tag,
                'equipment_name': eq.name if eq else tag,
                # Nombre comun para mostrar al mecanico (combina nombre + tag entre parentesis)
                'equipment_display': (f"{eq.name} ({tag})" if eq else tag),
                'activity': activity_label,
                'verb': verb_key,
                'responsible_party': row_party,
                'total_hours': round(total_hours, 2),
                'days': {d: {'label': day_labels.get(d, ''),
                             'item_ids': day_items.get(d, []),
                             'hours': day_hours.get(d, 0),
                             'status': day_status.get(d, '')} for d in range(7)},
            })

        rows.sort(key=lambda r: (r['area_name'], r['line_name'], r['equipment_name'], r['activity']))

        # Metadata de dias
        day_names_short = ['LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'DOM']
        days_meta = []
        try:
            base = dt.date.fromisoformat(plan.week_start)
            for i in range(7):
                d = base + dt.timedelta(days=i)
                days_meta.append({'idx': i, 'name': day_names_short[i],
                                  'date': d.isoformat(), 'is_weekend': i >= 5})
        except Exception:
            for i in range(7):
                days_meta.append({'idx': i, 'name': day_names_short[i],
                                  'date': '', 'is_weekend': i >= 5})

        return rows, days_meta

    @app.route('/api/weekly-plans/<int:plan_id>/matrix', methods=['GET'])
    def wp_matrix(plan_id):
        """Devuelve la matriz consolidada (tipo Excel del proveedor) en JSON."""
        try:
            p = WeeklyPlan.query.get_or_404(plan_id)
            rows, days_meta = _build_matrix(p)
            return jsonify({
                'plan': p.to_dict(),
                'days': days_meta,
                'rows': rows,
                'totals': {
                    'rows': len(rows),
                    'total_hours': round(sum(r['total_hours'] for r in rows), 2),
                    'executed_rows': sum(1 for r in rows if all(
                        d['status'] == 'EJECUTADO' or not d['item_ids']
                        for d in r['days'].values())),
                },
            })
        except Exception as e:
            logger.exception(f"wp_matrix error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans/<int:plan_id>/matrix/excel', methods=['GET'])
    def wp_matrix_excel(plan_id):
        """Exporta la matriz a XLSX con celdas P/✓ por dia (replica del Excel
        que el usuario envia al proveedor)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            p = WeeklyPlan.query.get_or_404(plan_id)
            rows, days_meta = _build_matrix(p)

            wb = Workbook()
            ws = wb.active
            ws.title = "Plan Semanal"

            # Estilos
            header_fill = PatternFill('solid', fgColor='1F4E79')
            header_font = Font(bold=True, color='FFFFFF', size=10)
            day_fill_lab = PatternFill('solid', fgColor='2E75B6')
            day_fill_wknd = PatternFill('solid', fgColor='C00000')
            mark_planned = PatternFill('solid', fgColor='2E75B6')
            mark_executed = PatternFill('solid', fgColor='70AD47')
            thin = Side(style='thin', color='BFBFBF')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Cabecera del documento
            ws.merge_cells('A1:K1')
            ws['A1'] = f"PROGRAMA SEMANAL DE MANTENIMIENTO - {p.code or 'PN'}"
            ws['A1'].font = Font(bold=True, size=14, color='1F4E79')
            ws['A1'].alignment = center

            ws.merge_cells('A2:K2')
            ws['A2'] = (f"Semana {p.week_start} al {p.week_end} · "
                        f"Proveedor: {p.provider.name if p.provider else 'Por asignar'} · "
                        f"Capacidad: {p.tech_count}t × {p.hours_per_night}h/noche")
            ws['A2'].alignment = center
            ws['A2'].font = Font(italic=True, size=9, color='5A6570')

            # Cabecera de columnas (fila 4) — ahora con LINEA y nombre comun de EQUIPO
            headers = ['AREA', 'LINEA', 'EQUIPO', 'ACTIVIDAD', 'HRS']
            for d in days_meta:
                headers.append(f"{d['name']}\n{d['date'][-5:] if d['date'] else ''}")
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=col_idx, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = center
                c.border = border
            # Marcar fines de semana (offset +1 por la nueva columna LINEA)
            for d in days_meta:
                if d['is_weekend']:
                    col = 6 + d['idx']
                    ws.cell(row=4, column=col).fill = day_fill_wknd

            # Filas de datos
            for r_idx, row in enumerate(rows, start=5):
                ws.cell(row=r_idx, column=1, value=row['area_name']).alignment = left
                ws.cell(row=r_idx, column=2, value=row['line_name']).alignment = left
                # Equipo: nombre comun primero, codigo entre parentesis (mecanico se guia por nombre)
                ws.cell(row=r_idx, column=3, value=row['equipment_display']).alignment = left
                ws.cell(row=r_idx, column=4, value=row['activity']).alignment = left
                ws.cell(row=r_idx, column=5, value=row['total_hours']).alignment = center
                for d in range(7):
                    cell = ws.cell(row=r_idx, column=6 + d)
                    day_data = row['days'][d]
                    if day_data['item_ids']:
                        if day_data['status'] == 'EJECUTADO':
                            cell.value = '✓'
                            cell.fill = mark_executed
                            cell.font = Font(bold=True, color='FFFFFF', size=11)
                        elif day_data['status'] == 'PARCIAL':
                            cell.value = '½'
                            cell.fill = PatternFill('solid', fgColor='FFC000')
                            cell.font = Font(bold=True, color='FFFFFF', size=11)
                        else:
                            cell.value = 'P'
                            cell.fill = mark_planned
                            cell.font = Font(bold=True, color='FFFFFF', size=11)
                    cell.alignment = center
                # Bordes a toda la fila (12 cols ahora)
                for col_idx in range(1, 13):
                    ws.cell(row=r_idx, column=col_idx).border = border

            # Anchos de columna
            ws.column_dimensions['A'].width = 16  # AREA
            ws.column_dimensions['B'].width = 18  # LINEA
            ws.column_dimensions['C'].width = 30  # EQUIPO (nombre + tag)
            ws.column_dimensions['D'].width = 42  # ACTIVIDAD
            ws.column_dimensions['E'].width = 8   # HRS
            for d in range(7):
                ws.column_dimensions[get_column_letter(6 + d)].width = 9

            # Congelar paneles (titulos siempre visibles)
            ws.freeze_panes = 'F5'

            # Pie con leyenda
            footer_row = 5 + len(rows) + 2
            ws.cell(row=footer_row, column=1,
                    value="Leyenda: P=Planificado · ½=Parcialmente ejecutado · ✓=Ejecutado").font = Font(italic=True, size=9)
            # Total de horas y filas
            total_h = sum(r['total_hours'] for r in rows)
            ws.cell(row=footer_row + 1, column=1,
                    value=f"Total: {len(rows)} actividades agrupadas · {round(total_h, 2)} horas planificadas").font = Font(italic=True, size=9, color='1F4E79')

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            filename = f"Plan_Semanal_{p.code or p.id}.xlsx"
            return send_file(
                bio, as_attachment=True, download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        except Exception as e:
            logger.exception(f"wp_matrix_excel error: {e}")
            return jsonify({"error": str(e)}), 500

    @app.route('/api/weekly-plans/<int:plan_id>/matrix/pdf', methods=['GET'])
    def wp_matrix_pdf(plan_id):
        """Exporta la matriz a PDF horizontal (apaisado) con celdas P/✓."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            )

            p = WeeklyPlan.query.get_or_404(plan_id)
            rows, days_meta = _build_matrix(p)

            bio = BytesIO()
            doc = SimpleDocTemplate(
                bio, pagesize=landscape(A4),
                leftMargin=10 * mm, rightMargin=10 * mm,
                topMargin=10 * mm, bottomMargin=10 * mm,
                title=f"Plan Semanal {p.code or p.id}",
            )
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('t', parent=styles['Title'], fontSize=14,
                                         textColor=colors.HexColor('#1F4E79'), alignment=1)
            sub_style = ParagraphStyle('s', parent=styles['Normal'], fontSize=9,
                                       textColor=colors.HexColor('#5A6570'), alignment=1)
            cell_style = ParagraphStyle('c', parent=styles['Normal'], fontSize=7, leading=8)
            cell_bold = ParagraphStyle('cb', parent=styles['Normal'], fontSize=7, leading=8,
                                       fontName='Helvetica-Bold')

            story = []
            story.append(Paragraph(f"PROGRAMA SEMANAL DE MANTENIMIENTO — {p.code or 'PN'}", title_style))
            story.append(Paragraph(
                f"Semana {p.week_start} al {p.week_end} · "
                f"Proveedor: {p.provider.name if p.provider else 'Por asignar'} · "
                f"{p.tech_count} técnicos × {p.hours_per_night}h/noche",
                sub_style))
            story.append(Spacer(1, 4 * mm))

            # Cabecera de tabla — ahora con LINEA
            header_row = ['ÁREA', 'LÍNEA', 'EQUIPO', 'ACTIVIDAD', 'HRS']
            for d in days_meta:
                header_row.append(f"{d['name']}\n{d['date'][-5:] if d['date'] else ''}")

            table_data = [header_row]
            for r in rows:
                row_cells = [
                    Paragraph(r['area_name'], cell_style),
                    Paragraph(r['line_name'], cell_style),
                    # Equipo: nombre comun en negrita, codigo abajo en gris
                    Paragraph(f"<b>{r['equipment_name']}</b><br/><font color='#888888' size='6'>{r['equipment_tag']}</font>", cell_style),
                    Paragraph(r['activity'], cell_style),
                    f"{r['total_hours']}h",
                ]
                for d in range(7):
                    dd = r['days'][d]
                    if dd['item_ids']:
                        if dd['status'] == 'EJECUTADO':
                            row_cells.append('✓')
                        elif dd['status'] == 'PARCIAL':
                            row_cells.append('½')
                        else:
                            row_cells.append('P')
                    else:
                        row_cells.append('')
                table_data.append(row_cells)

            tbl = Table(
                table_data,
                colWidths=[18 * mm, 22 * mm, 38 * mm, 60 * mm, 10 * mm,
                           14 * mm, 14 * mm, 14 * mm, 14 * mm, 14 * mm, 14 * mm, 14 * mm],
                repeatRows=1,
            )
            ts = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 8),
                ('FONT', (0, 1), (-1, -1), 'Helvetica', 7),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#BFBFBF')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (4, 0), (-1, -1), 'CENTER'),
                ('PADDING', (0, 0), (-1, -1), 2),
            ])
            # Pintar fines de semana en cabecera (offset +1 por LINEA)
            for d in days_meta:
                if d['is_weekend']:
                    col = 5 + d['idx']
                    ts.add('BACKGROUND', (col, 0), (col, 0), colors.HexColor('#C00000'))
            # Pintar celdas marcadas
            for r_idx, r in enumerate(rows, start=1):
                for d in range(7):
                    dd = r['days'][d]
                    if dd['item_ids']:
                        col = 5 + d
                        if dd['status'] == 'EJECUTADO':
                            ts.add('BACKGROUND', (col, r_idx), (col, r_idx), colors.HexColor('#70AD47'))
                            ts.add('TEXTCOLOR', (col, r_idx), (col, r_idx), colors.white)
                        elif dd['status'] == 'PARCIAL':
                            ts.add('BACKGROUND', (col, r_idx), (col, r_idx), colors.HexColor('#FFC000'))
                            ts.add('TEXTCOLOR', (col, r_idx), (col, r_idx), colors.white)
                        else:
                            ts.add('BACKGROUND', (col, r_idx), (col, r_idx), colors.HexColor('#2E75B6'))
                            ts.add('TEXTCOLOR', (col, r_idx), (col, r_idx), colors.white)
                        ts.add('FONT', (col, r_idx), (col, r_idx), 'Helvetica-Bold', 9)

            tbl.setStyle(ts)
            story.append(tbl)
            story.append(Spacer(1, 3 * mm))
            story.append(Paragraph(
                "<i>Leyenda: P = Planificado · ½ = Parcialmente ejecutado · ✓ = Ejecutado</i>",
                ParagraphStyle('lg', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#5A6570'))))

            doc.build(story)
            bio.seek(0)
            filename = f"Plan_Semanal_{p.code or p.id}.pdf"
            return send_file(
                bio, as_attachment=True, download_name=filename,
                mimetype='application/pdf',
            )
        except Exception as e:
            logger.exception(f"wp_matrix_pdf error: {e}")
            return jsonify({"error": str(e)}), 500

    # ── Export PDF ───────────────────────────────────────────────────────

    @app.route('/api/weekly-plans/<int:plan_id>/report/pdf', methods=['GET'])
    def wp_export_pdf(plan_id):
        """PDF con 7 secciones (una por noche), listo para entregar al proveedor."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
            )

            p = WeeklyPlan.query.get_or_404(plan_id)

            bio = BytesIO()
            doc = SimpleDocTemplate(
                bio, pagesize=A4,
                leftMargin=12 * mm, rightMargin=12 * mm,
                topMargin=12 * mm, bottomMargin=12 * mm,
                title=f"Programa Nocturno {p.code or p.id}",
            )
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('t', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#0a84ff'), alignment=1)
            subtitle_style = ParagraphStyle('s', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#5a6570'), alignment=1)
            section_style = ParagraphStyle('sec', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#BF5AF2'), spaceBefore=6)
            cell_style = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=10)
            cell_bold = ParagraphStyle('cellb', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold')

            story = []
            story.append(Paragraph("PROGRAMA NOCTURNO SEMANAL", title_style))
            story.append(Paragraph(
                f"<b>{p.code or ''}</b> · Semana {p.week_start} al {p.week_end}"
                f" · Proveedor: {p.provider.name if p.provider else 'Por asignar'}"
                f" · Capacidad: {p.tech_count} técnicos × {p.hours_per_night}h = {p.tech_count * p.hours_per_night}h/noche",
                subtitle_style))
            story.append(Spacer(1, 5 * mm))

            # Una sección por día con ítems agrupados por área
            day_names = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO']
            capacity = p.tech_count * p.hours_per_night

            for d_idx in range(7):
                day_items = [it for it in p.items if it.day_of_week == d_idx]
                if not day_items:
                    continue
                day_hours = sum(it.estimated_hours for it in day_items)
                day_date = dt.date.fromisoformat(p.week_start) + dt.timedelta(days=d_idx)
                # Conteo por disciplina para cabecera del dia
                disc_counts = {'MECANICO': 0, 'ELECTRICO': 0, 'MIXTO': 0, 'SIN CLASIF': 0}
                items_with_disc = []
                for it in day_items:
                    d = discipline_for_weekly_item(it)
                    items_with_disc.append((it, d))
                    disc_counts[d] = disc_counts.get(d, 0) + 1

                breakdown = ' · '.join(
                    f"{k.title()}: {v}" for k, v in disc_counts.items() if v > 0
                )
                story.append(Paragraph(
                    f"🌙 {day_names[d_idx]} — {day_date.isoformat()} · "
                    f"{len(day_items)} tareas · {day_hours:.1f}h / {capacity:.0f}h · {breakdown}",
                    section_style))

                # Tabla
                rows = [['#', 'Área', 'Disciplina', 'Tipo', 'Código', 'Descripción', 'Hrs', 'Check']]
                items_with_disc.sort(key=lambda x: (x[0].area.name if x[0].area else '', x[0].order_index))
                disc_short = {'MECANICO': 'MEC', 'ELECTRICO': 'ELEC', 'MIXTO': 'MIX', 'SIN CLASIF': '-'}
                for i, (it, disc) in enumerate(items_with_disc, 1):
                    rows.append([
                        str(i),
                        Paragraph((it.area.name if it.area else '-'), cell_style),
                        disc_short.get(disc, disc),
                        (it.source_type or '').upper(),
                        Paragraph(it.source_code or '-', cell_bold),
                        Paragraph((it.description or it.source_name or '-')[:200], cell_style),
                        f"{it.estimated_hours}h",
                        '[  ]',
                    ])
                tbl = Table(
                    rows,
                    colWidths=[8 * mm, 26 * mm, 14 * mm, 18 * mm, 22 * mm, 75 * mm, 12 * mm, 10 * mm],
                    repeatRows=1,
                )
                tbl.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1b1d3a')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 8),
                    ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
                    ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
                ]))
                story.append(tbl)
                story.append(Spacer(1, 4 * mm))

            if p.notes:
                story.append(Paragraph("NOTAS", section_style))
                story.append(Paragraph(p.notes.replace('\n', '<br/>'), cell_style))

            doc.build(story)
            bio.seek(0)
            filename = f"Programa_Nocturno_{p.code or p.id}.pdf"
            return send_file(
                bio, as_attachment=True, download_name=filename,
                mimetype='application/pdf',
            )
        except Exception as e:
            logger.error(f"export_weekly_plan_pdf error: {e}")
            import traceback; traceback.print_exc()
            return jsonify({"error": str(e)}), 500
