"""Generate hybrid asset tree Excel proposal."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Arbol Hibrido Propuesto"

hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
rf = Font(bold=True, color="C0392B", size=10)
rfill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
tb = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

headers = ["Area", "Linea", "Equipo", "TagEquipo", "Sistema", "Componente", "Criticidad", "Nota"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center'); c.border = tb

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 28
ws.column_dimensions['F'].width = 32
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 55

dig_sys = {
    "SISTEMA DE ACCIONAMIENTO": [
        ("Motor Electrico", "Media", "ACTIVO ROTATIVO (marca, modelo, serial en ficha)"),
        ("Caja Reductora", "Media", "ACTIVO ROTATIVO (marca, modelo, serial en ficha)"),
        ("Acople Fusible", "Media", ""),
        ("Acople", "Media", ""),
        ("Brida", "Media", ""),
        ("Faja", "Media", "Pieza de desgaste - spec: tipo, medida"),
        ("Polea Motriz", "Media", ""),
        ("Polea Conducido", "Media", ""),
        ("Chumacera Motriz", "Media", ""),
        ("Chumacera Conducida", "Media", ""),
    ],
    "SISTEMA DE VAPOR": [
        ("Tuberia de Vapor Entrada 1", "Media", ""),
        ("Valvula de Ingreso de Vapor 1", "Media", ""),
        ("Manometro 1", "Media", ""),
        ("Tuberia de Vapor Entrada 2", "Media", ""),
        ("Valvula de Ingreso de Vapor 2", "Media", ""),
        ("Manometro 2", "Media", ""),
        ("Valvula de Condensado 1", "Media", ""),
        ("Valvula de Condensado 2", "Media", ""),
        ("Valvula Check 1", "Media", ""),
        ("Valvula Check 2", "Media", ""),
        ("Trampa de Vapor 1", "Media", ""),
        ("Trampa de Vapor 2", "Media", ""),
        ("Filtro Y 1", "Media", ""),
        ("Filtro Y 2", "Media", ""),
        ("Manometro de Chaqueta", "Media", ""),
        ("Manometro de Descarga", "Media", ""),
        ("Manometro de Salida", "Media", ""),
        ("Ducto Descarga de Vapores", "Media", ""),
        ("Valvula Descarga de Vapores", "Media", ""),
    ],
    "TANQUE DIGESTOR": [
        ("Boca de Llenado", "Media", ""),
        ("Chaqueta Interna", "Media", ""),
        ("Chaqueta Exterior", "Media", ""),
        ("Enchaquetado Exterior", "Media", ""),
        ("Eje Lado Conducido", "Media", ""),
        ("Eje Lado Motriz", "Media", ""),
        ("Escotilla de Descarga", "Media", ""),
        ("Prensa Estopa Lado Conducido", "Media", ""),
        ("Prensa Estopa Lado Motriz", "Media", ""),
        ("Tapa Bombeada Conducido", "Media", ""),
        ("Tapa Bombeada Motriz", "Media", ""),
        ("Tripode", "Media", ""),
    ],
    "SISTEMA DE SOPORTE": [
        ("Placa Base", "Media", ""),
    ],
    "AUXILIARES": [
        ("Guarda de Faja", "Media", ""),
        ("Guarda Motor Electrico", "Media", ""),
    ],
}

th_sys = {
    "SISTEMA DE ACCIONAMIENTO": [
        ("Motorreductor", "Media", "ACTIVO ROTATIVO (marca, modelo, RPM, potencia en ficha)"),
        ("Cadena", "Media", "Pieza de desgaste - spec: tipo, paso, eslabones"),
        ("Sprocket Motriz", "Media", "Spec: dientes, paso"),
        ("Sprocket Conducido", "Media", "Spec: dientes, paso"),
        ("Chumacera Motriz", "Media", ""),
        ("Chumacera Conducida", "Media", ""),
        ("Chaveta", "Media", ""),
    ],
    "TORNILLO SINFIN": [
        ("Tubo Central", "Media", ""),
        ("Helice", "Media", "Spec: diametro, paso"),
        ("Puente", "Media", "Spec: plano bocina bronce"),
        ("Eje Motriz", "Media", ""),
        ("Eje de Cola", "Media", ""),
        ("Eje Central", "Media", ""),
    ],
    "SISTEMA CUERPO": [
        ("Tina", "Media", ""),
        ("Tapas Superiores", "Media", ""),
        ("Tapas Laterales", "Media", ""),
        ("Chute de Alimentacion", "Media", ""),
        ("Chute de Descarga", "Media", ""),
        ("Chute Central", "Media", ""),
        ("Patin", "Media", ""),
    ],
    "SISTEMA SOPORTE": [
        ("Placa Base de Motorreductor", "Media", ""),
        ("Soporte de Tina", "Media", ""),
        ("Pluma de Izaje", "Media", ""),
    ],
    "AUXILIARES": [
        ("Guarda Cadena", "Media", ""),
        ("Guarda para Aceite", "Media", ""),
        ("Guarda para Chumacera", "Media", ""),
        ("Guarda para Faja", "Media", ""),
    ],
}

other = [
    ("CALDERAS", "CALDERA 400 BHP", "", ""),
    ("CALDERAS", "CALDERA 900 BHP", "", ""),
    ("CALDERAS", "MANIFOLD DE VAPOR", "", ""),
    ("CALDERAS", "OSMOSIS", "", ""),
    ("COCCION", "HIDROLAVADORAS", "", ""),
    ("COCCION", "LINEA PERCOLADOR", "PERCOLADOR #1", "PER1"),
    ("COCCION", "LINEA PERCOLADOR", "PERCOLADOR #2", "PER2"),
    ("LINEA DE POLLO", "", "", ""),
    ("MOLINO", "CICLON DE ENSAQUE", "", ""),
    ("MOLINO", "CICLON DE LLEGADA", "", ""),
    ("MOLINO", "FAJA TRANSPORTADORA", "", ""),
    ("MOLINO", "LINEA MOLINO #1", "", ""),
    ("MOLINO", "LINEA MOLINO #2", "", ""),
    ("MOLINO", "LINEA ZARANDA", "", ""),
    ("RMP", "HIDROLAVADORAS", "HIDROLAVADORA #1", "H1"),
    ("RMP", "HIDROLAVADORAS", "HIDROLAVADORA #2", "H2"),
    ("RMP", "HIDROLAVADORAS", "HIDROLAVADORA #3", "H3"),
    ("SECADO", "ENFRIADOR", "", ""),
    ("SECADO", "LANZAHARINA", "", ""),
    ("SECADO", "PURIFICADOR", "", ""),
    ("SECADO", "SECADOR #1", "", ""),
    ("SECADO", "SECADOR #2", "", ""),
    ("SUBESTACION ELECTRICA", "", "", ""),
    ("TRITURADO", "HIDROLAVADORA", "HIDROLAVADORA #5", "H5"),
    ("TRITURADO", "TRITURADOR GRANDE", "TH ALIMENTADOR", "THALI"),
    ("TRITURADO", "TRITURADOR GRANDE", "TH SALIDA", "THSAL"),
    ("TRITURADO", "TRITURADOR GRANDE", "TH SILO", "THSI"),
    ("TRITURADO", "TRITURADOR GRANDE", "TRITURADOR 100 HP", "TRI1"),
    ("TRITURADO", "TRITURADOR PEQUENO", "TH SALIDA", "THSAL"),
    ("TRITURADO", "TRITURADOR PEQUENO", "TRITURADOR 75 HP", "TRI2"),
    ("UTILITIES", "", "", ""),
    ("VAHOS", "HIDROLISIS", "", ""),
    ("VAHOS", "VAHOS #1", "", ""),
    ("VAHOS", "VAHOS #2", "", ""),
]

row = 2

def wr(r, a, l, e, t, s, co, cr, n, is_rot=False):
    vals = [a, l, e, t, s, co, cr, n]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = tb
        cell.alignment = Alignment(vertical='center')
        if is_rot:
            cell.font = rf
            cell.fill = rfill

for n in range(1, 10):
    for sys_name, comps in dig_sys.items():
        for comp_name, crit, nota in comps:
            is_rot = "ACTIVO ROTATIVO" in nota
            wr(row, "COCCION", f"LINEA DIGESTOR #{n}", f"DIGESTOR #{n}", f"D{n}", sys_name, comp_name, crit, nota, is_rot)
            row += 1
    for sys_name, comps in th_sys.items():
        for comp_name, crit, nota in comps:
            is_rot = "ACTIVO ROTATIVO" in nota
            wr(row, "COCCION", f"LINEA DIGESTOR #{n}", f"TH{n}", f"TH{n}", sys_name, comp_name, crit, nota, is_rot)
            row += 1

for a, l, e, t in other:
    wr(row, a, l, e, t, "", "", "", "")
    row += 1

# Sheet 2: Activos Rotativos
ws2 = wb.create_sheet("Activos Rotativos Sugeridos")
h2 = ["Codigo", "Nombre", "Categoria", "Marca", "Modelo", "Serial", "RPM", "Potencia", "Ubicacion", "Specs en ficha"]
for c, h in enumerate(h2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = hfill; cell.border = tb
ws2.column_dimensions['A'].width = 14; ws2.column_dimensions['B'].width = 28
ws2.column_dimensions['C'].width = 18; ws2.column_dimensions['I'].width = 35; ws2.column_dimensions['J'].width = 60

r2 = 2
for n in range(1, 10):
    for code, name, cat, ubi, specs in [
        (f"MOT-D{n}", f"Motor Electrico Digestor #{n}", "Motor Electrico", f"DIGESTOR #{n} [D{n}] > Motor Electrico",
         "Rodamiento motriz, Rodamiento conducido, Estator, Rotor, Ventilador, Carcasa, Grasa, Eje salida, Cubierta ventilador"),
        (f"RED-D{n}", f"Caja Reductora Digestor #{n}", "Caja Reductora", f"DIGESTOR #{n} [D{n}] > Caja Reductora",
         "Rodamiento entrada, Rodamiento salida, Reten entrada, Reten salida, Engranaje, Eje entrada, Eje salida, Aceite, Visor aceite"),
        (f"MR-TH{n}", f"Motorreductor TH{n}", "Motorreductor", f"TH{n} [TH{n}] > Motorreductor",
         "Motor electrico, Rodamiento motriz, Rodamiento conducido, Reten central, Reten salida, Eje salida, Aceite"),
    ]:
        ws2.cell(row=r2, column=1, value=code).border = tb
        ws2.cell(row=r2, column=2, value=name).border = tb
        ws2.cell(row=r2, column=3, value=cat).border = tb
        ws2.cell(row=r2, column=4, value="(completar)").border = tb
        ws2.cell(row=r2, column=5, value="(completar)").border = tb
        ws2.cell(row=r2, column=6, value="(completar)").border = tb
        ws2.cell(row=r2, column=7, value="(completar)").border = tb
        ws2.cell(row=r2, column=8, value="(completar)").border = tb
        ws2.cell(row=r2, column=9, value=ubi).border = tb
        ws2.cell(row=r2, column=10, value=specs).border = tb
        r2 += 1

# Sheet 3: Eliminados
ws3 = wb.create_sheet("Componentes Eliminados")
h3 = ["Equipo", "Sistema Original", "Componente Eliminado", "Ahora va en", "Justificacion"]
for c, h in enumerate(h3, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = hf; cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid"); cell.border = tb
ws3.column_dimensions['A'].width = 16; ws3.column_dimensions['B'].width = 20; ws3.column_dimensions['C'].width = 28
ws3.column_dimensions['D'].width = 35; ws3.column_dimensions['E'].width = 45

r3 = 2
elim = [
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Rodamiento Motriz", "Ficha MOT-Dx", "Parte interna del motor"),
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Rodamiento Conducido", "Ficha MOT-Dx", "Parte interna del motor"),
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Estator", "Ficha MOT-Dx", "Parte interna del motor"),
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Rotor", "Ficha MOT-Dx", "Parte interna del motor"),
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Ventilador", "Ficha MOT-Dx", "Parte interna del motor"),
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Carcasa", "Ficha MOT-Dx", "Estructura del motor"),
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Cubierta Ventilador", "Ficha MOT-Dx", "Accesorio del motor"),
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Grasa", "Ficha MOT-Dx", "Consumible del motor"),
    ("DIGESTOR #x", "MOTOR ELECTRICO", "Eje Salida", "Ficha MOT-Dx", "Parte interna del motor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Rodamiento Entrada", "Ficha RED-Dx", "Parte interna del reductor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Rodamiento Salida", "Ficha RED-Dx", "Parte interna del reductor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Reten Entrada", "Ficha RED-Dx", "Pieza desgaste del reductor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Reten Salida", "Ficha RED-Dx", "Pieza desgaste del reductor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Engranaje", "Ficha RED-Dx", "Parte interna del reductor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Eje de Entrada", "Ficha RED-Dx", "Parte interna del reductor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Eje de Salida", "Ficha RED-Dx", "Parte interna del reductor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Aceite", "Ficha RED-Dx", "Consumible del reductor"),
    ("DIGESTOR #x", "CAJA REDUCTORA", "Visor de Aceite", "Ficha RED-Dx", "Accesorio del reductor"),
    ("THx", "MOTORREDUCTOR", "Motor Electrico", "Ficha MR-THx", "Parte integrada del motorreductor"),
    ("THx", "MOTORREDUCTOR", "Rodamiento Motriz", "Ficha MR-THx", "Parte interna"),
    ("THx", "MOTORREDUCTOR", "Rodamiento Conducido", "Ficha MR-THx", "Parte interna"),
    ("THx", "MOTORREDUCTOR", "Reten Central", "Ficha MR-THx", "Pieza desgaste"),
    ("THx", "MOTORREDUCTOR", "Reten de Salida", "Ficha MR-THx", "Pieza desgaste"),
    ("THx", "MOTORREDUCTOR", "Eje de Salida", "Ficha MR-THx", "Parte interna"),
    ("THx", "MOTORREDUCTOR", "Aceite", "Ficha MR-THx", "Consumible"),
]
for eq, sys, comp, dest, just in elim:
    ws3.cell(row=r3, column=1, value=eq).border = tb
    ws3.cell(row=r3, column=2, value=sys).border = tb
    ws3.cell(row=r3, column=3, value=comp).border = tb
    ws3.cell(row=r3, column=4, value=dest).border = tb
    ws3.cell(row=r3, column=5, value=just).border = tb
    r3 += 1

wb.save("d:/PROGRAMACION/CMMS_Industrial/arbol_hibrido_propuesto.xlsx")
print("OK: arbol_hibrido_propuesto.xlsx generado")
