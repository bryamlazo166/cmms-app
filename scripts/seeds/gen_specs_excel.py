"""Genera un Excel con todos los equipos del arbol y specs tipicas prellenadas
con valores comunes de la industria (editable antes de cargar).

Uso: python gen_specs_excel.py
Salida: specs_equipos_template.xlsx
"""
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app
from database import db
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Columnas generales + sugerencia por tipo
SPEC_COLUMNS = [
    "AREA", "LINEA", "TAG", "EQUIPO", "TIPO",
    "MARCA", "MODELO", "N_SERIE", "ANO_FABRICACION",
    "POTENCIA_HP", "POTENCIA_KW", "TENSION_V", "CORRIENTE_NOMINAL_A", "RPM",
    "TIPO_ARRANQUE",
    "PROTECCION_IP", "CLASE_AISLAMIENTO",
    "PESO_KG", "LARGO_M", "ANCHO_M", "ALTO_M",
    "CAPACIDAD_PRODUCCION", "UNIDAD_CAPACIDAD",
    "TEMPERATURA_OPERACION_C", "PRESION_OPERACION_BAR",
    "LUBRICANTE_PRINCIPAL", "LUBRICANTE_REDUCTOR",
    "CRITICIDAD", "OBSERVACIONES",
]

# Valores sugeridos por tipo de equipo (industria alimentos/bio-mass típica)
DEFAULTS = {
    "DIGESTOR": {
        "TIPO": "Digestor cocción",
        "MARCA": "FABRICACION LOCAL",
        "POTENCIA_HP": 15, "POTENCIA_KW": 11, "TENSION_V": 440,
        "CORRIENTE_NOMINAL_A": 18, "RPM": 1750,
        "TIPO_ARRANQUE": "Directo (DOL)",
        "PROTECCION_IP": "IP55", "CLASE_AISLAMIENTO": "F",
        "CAPACIDAD_PRODUCCION": 8, "UNIDAD_CAPACIDAD": "TM/h",
        "TEMPERATURA_OPERACION_C": 120, "PRESION_OPERACION_BAR": 3,
        "LUBRICANTE_PRINCIPAL": "GRASA FRIXO 177",
        "LUBRICANTE_REDUCTOR": "ACEITE ISO VG 320",
        "CRITICIDAD": "Alta",
    },
    "TH": {  # Transportadores helicoidales
        "TIPO": "Transportador helicoidal",
        "MARCA": "FABRICACION LOCAL",
        "POTENCIA_HP": 7.5, "POTENCIA_KW": 5.5, "TENSION_V": 440,
        "CORRIENTE_NOMINAL_A": 9, "RPM": 1750,
        "TIPO_ARRANQUE": "Directo (DOL)",
        "PROTECCION_IP": "IP55", "CLASE_AISLAMIENTO": "F",
        "CAPACIDAD_PRODUCCION": 5, "UNIDAD_CAPACIDAD": "TM/h",
        "LUBRICANTE_PRINCIPAL": "GRASA FRIXO 177",
        "LUBRICANTE_REDUCTOR": "ACEITE ISO VG 320",
        "CRITICIDAD": "Media",
    },
    "CICLON": {
        "TIPO": "Ciclón separador",
        "MARCA": "FABRICACION LOCAL",
        "CAPACIDAD_PRODUCCION": 2000, "UNIDAD_CAPACIDAD": "m3/h aire",
        "CRITICIDAD": "Media",
    },
    "MOLINO": {
        "TIPO": "Molino martillos",
        "MARCA": "BLISS / PRATER",
        "POTENCIA_HP": 150, "POTENCIA_KW": 110, "TENSION_V": 440,
        "CORRIENTE_NOMINAL_A": 180, "RPM": 3600,
        "TIPO_ARRANQUE": "Estrella-Triangulo",
        "PROTECCION_IP": "IP55", "CLASE_AISLAMIENTO": "F",
        "CAPACIDAD_PRODUCCION": 3, "UNIDAD_CAPACIDAD": "TM/h",
        "LUBRICANTE_PRINCIPAL": "GRASA NLGI 2",
        "CRITICIDAD": "Alta",
    },
    "PERCOLADOR": {
        "TIPO": "Percolador",
        "POTENCIA_HP": 10, "POTENCIA_KW": 7.5, "TENSION_V": 440,
        "TIPO_ARRANQUE": "Directo (DOL)",
        "CRITICIDAD": "Media",
    },
    "HIDROLAVADORA": {
        "TIPO": "Hidrolavadora alta presion",
        "MARCA": "KARCHER / HIDROMAC",
        "POTENCIA_HP": 15, "POTENCIA_KW": 11, "TENSION_V": 440,
        "CORRIENTE_NOMINAL_A": 18, "RPM": 1450,
        "TIPO_ARRANQUE": "Directo (DOL)",
        "PROTECCION_IP": "IP55",
        "PRESION_OPERACION_BAR": 200,
        "CAPACIDAD_PRODUCCION": 20, "UNIDAD_CAPACIDAD": "L/min",
        "CRITICIDAD": "Media",
    },
    "SECADOR": {
        "TIPO": "Secador rotativo",
        "POTENCIA_HP": 40, "POTENCIA_KW": 30, "TENSION_V": 440,
        "TIPO_ARRANQUE": "Variador (VFD)",
        "CAPACIDAD_PRODUCCION": 4, "UNIDAD_CAPACIDAD": "TM/h",
        "TEMPERATURA_OPERACION_C": 180,
        "LUBRICANTE_REDUCTOR": "ACEITE ISO VG 320",
        "CRITICIDAD": "Alta",
    },
    "PURIFICADOR": {
        "TIPO": "Purificador",
        "POTENCIA_HP": 10, "POTENCIA_KW": 7.5, "TENSION_V": 440,
        "TIPO_ARRANQUE": "Directo (DOL)",
        "CRITICIDAD": "Media",
    },
    "TRITURADOR": {
        "TIPO": "Triturador",
        "MARCA": "FABRICACION LOCAL",
        "POTENCIA_HP": 100, "POTENCIA_KW": 75, "TENSION_V": 440,
        "CORRIENTE_NOMINAL_A": 120, "RPM": 1750,
        "TIPO_ARRANQUE": "Estrella-Triangulo",
        "PROTECCION_IP": "IP55", "CLASE_AISLAMIENTO": "F",
        "CAPACIDAD_PRODUCCION": 4, "UNIDAD_CAPACIDAD": "TM/h",
        "LUBRICANTE_PRINCIPAL": "GRASA NLGI 2",
        "CRITICIDAD": "Alta",
    },
    "FAJA": {
        "TIPO": "Faja transportadora",
        "POTENCIA_HP": 7.5, "POTENCIA_KW": 5.5, "TENSION_V": 440,
        "TIPO_ARRANQUE": "Directo (DOL)",
        "CAPACIDAD_PRODUCCION": 10, "UNIDAD_CAPACIDAD": "TM/h",
        "CRITICIDAD": "Media",
    },
}


def infer_type(tag, name):
    u = (name or '').upper() + ' ' + (tag or '').upper()
    if 'DIGESTOR' in u or tag.upper().startswith('D') and tag.upper()[1:].isdigit():
        if 'DIGESTOR' in u:
            return 'DIGESTOR'
    if tag.upper().startswith('TH') or 'TH ' in u or u.startswith('TH'):
        return 'TH'
    if 'CICLON' in u:
        return 'CICLON'
    if 'MOLINO' in u:
        return 'MOLINO'
    if 'PERCOLADOR' in u:
        return 'PERCOLADOR'
    if 'HIDROLAV' in u:
        return 'HIDROLAVADORA'
    if 'SECADOR' in u:
        return 'SECADOR'
    if 'PURIFIC' in u:
        return 'PURIFICADOR'
    if 'TRITURA' in u:
        return 'TRITURADOR'
    if 'FAJA' in u:
        return 'FAJA'
    if 'DIGESTOR' in u:
        return 'DIGESTOR'
    return 'TH'  # default seguro para transportadores no tipificados


def run():
    with app.app_context():
        rows = db.session.execute(text("""
            SELECT e.tag, e.name, l.name AS linea, a.name AS area
            FROM equipments e
            LEFT JOIN lines l ON e.line_id = l.id
            LEFT JOIN areas a ON l.area_id = a.id
            ORDER BY a.name, l.name, e.name
        """)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Specs Equipos"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(SPEC_COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for r in rows:
        tag, name, linea, area = r
        t = infer_type(tag, name)
        defaults = DEFAULTS.get(t, {})
        row = [
            area or '', linea or '', tag or '', name or '',
            defaults.get("TIPO", ''),
            defaults.get("MARCA", ''), '', '', '',
            defaults.get("POTENCIA_HP", ''), defaults.get("POTENCIA_KW", ''),
            defaults.get("TENSION_V", ''), defaults.get("CORRIENTE_NOMINAL_A", ''),
            defaults.get("RPM", ''),
            defaults.get("TIPO_ARRANQUE", ''),
            defaults.get("PROTECCION_IP", ''), defaults.get("CLASE_AISLAMIENTO", ''),
            '', '', '', '',
            defaults.get("CAPACIDAD_PRODUCCION", ''), defaults.get("UNIDAD_CAPACIDAD", ''),
            defaults.get("TEMPERATURA_OPERACION_C", ''), defaults.get("PRESION_OPERACION_BAR", ''),
            defaults.get("LUBRICANTE_PRINCIPAL", ''), defaults.get("LUBRICANTE_REDUCTOR", ''),
            defaults.get("CRITICIDAD", 'Media'), '',
        ]
        ws.append(row)

    widths = {
        'A': 14, 'B': 22, 'C': 12, 'D': 28, 'E': 24, 'F': 18, 'G': 16, 'H': 14,
        'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 12, 'N': 10, 'O': 20,
        'P': 12, 'Q': 14, 'R': 10, 'S': 10, 'T': 10, 'U': 10, 'V': 16, 'W': 14,
        'X': 16, 'Y': 16, 'Z': 24, 'AA': 22, 'AB': 10, 'AC': 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'E2'

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instr = [
        ["PLANTILLA DE ESPECIFICACIONES TECNICAS"],
        [f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        [""],
        ["Columnas clave a completar:"],
        ["- MARCA, MODELO, N_SERIE, ANO_FABRICACION: obtener de placa del equipo"],
        ["- POTENCIA, TENSION, CORRIENTE, RPM: de placa del motor principal"],
        ["- TIPO_ARRANQUE: Directo (DOL) | Estrella-Triangulo | Softstarter | Variador (VFD)"],
        ["- PROTECCION_IP: IP54, IP55, IP65"],
        ["- CLASE_AISLAMIENTO: B, F, H"],
        ["- CRITICIDAD: Alta | Media | Baja"],
        [""],
        ["Los valores prellenados son SUGERENCIAS basadas en tipicos de industria."],
        ["Reemplaza con los datos reales de cada equipo."],
        [""],
        ["Cuando termines de llenar, avisame y te creo el script que carga estos"],
        ["datos al sistema (tabla equipment_specs) automaticamente."],
    ]
    for row in instr:
        ws2.append(row)
    ws2.column_dimensions['A'].width = 90
    ws2['A1'].font = Font(bold=True, size=14)

    out = os.path.join(os.path.dirname(__file__), 'specs_equipos_template.xlsx')
    wb.save(out)
    print(f"Excel generado: {out}")
    print(f"Equipos incluidos: {len(rows)}")


if __name__ == '__main__':
    run()
