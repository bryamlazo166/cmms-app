"""Genera Excel con Plan de Mantenimiento Preventivo (Inspecciones + Lubricaciones)
para entregar al proveedor de turno noche.

Areas: COCCION, MOLINO, SECADO (sin Hidrolavadoras)
Frecuencias: Molinos=interdiaria, resto=semanal/quincenal

Uso: python gen_plan_preventivo.py
Salida: plan_preventivo_proveedor.xlsx
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# ── Tareas por tipo de equipo ──────────────────────────────────────────────────

def _tareas_digestor(tag, name):
    """D1-D9: semanal y quincenal"""
    return [
        # LUBRICACION
        (tag, name, "COCCION", "LUB", "Lubricación chumacera motriz", "Engrasar chumacera lado motriz con pistola manual. Verificar temperatura post-engrase.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "COCCION", "LUB", "Lubricación chumacera conducida", "Engrasar chumacera lado conducido con pistola manual. Verificar temperatura post-engrase.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "COCCION", "LUB", "Lubricación cadena de transmisión", "Aplicar lubricante en cadena de transmisión. Verificar tensión y desgaste de eslabones.", "CRC BELT GRIP / ACEITE CADENA", "QUINCENAL", 15),
        (tag, name, "COCCION", "LUB", "Nivel aceite caja reductora", "Verificar nivel de aceite en visor de caja reductora. Completar si está por debajo del mínimo.", "ACEITE ISO VG 320", "QUINCENAL", 15),
        # INSPECCION
        (tag, name, "COCCION", "INSP", "Inspección fajas de transmisión", "Verificar estado, tensión, alineación y desgaste de fajas. Revisar grietas, deshilachado o patinaje.", "-", "SEMANAL", 7),
        (tag, name, "COCCION", "INSP", "Inspección poleas motriz y conducida", "Verificar desgaste de canales, alineación entre poleas, fisuras. Usar regla de alineación.", "-", "QUINCENAL", 15),
        (tag, name, "COCCION", "INSP", "Inspección motor eléctrico", "Verificar temperatura, ruido anormal, vibración, estado de ventilador, caja de bornes y cableado.", "-", "SEMANAL", 7),
        (tag, name, "COCCION", "INSP", "Inspección guardas de seguridad", "Verificar que guardas de faja y motor estén en posición, completas y bien fijadas.", "-", "QUINCENAL", 15),
        (tag, name, "COCCION", "INSP", "Inspección fugas de vapor", "Revisar uniones, válvulas, trampas de vapor, empaquetaduras. Reportar fugas.", "-", "SEMANAL", 7),
        (tag, name, "COCCION", "INSP", "Inspección manómetros", "Verificar lectura correcta de manómetros de chaqueta, descarga y salida. Reportar si están dañados o sin lectura.", "-", "QUINCENAL", 15),
    ]


def _tareas_th_coccion(tag, name):
    """TH1-TH9 de digestores: semanal y quincenal"""
    return [
        (tag, name, "COCCION", "LUB", "Lubricación chumacera motriz", "Engrasar chumacera lado motriz.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "COCCION", "LUB", "Lubricación chumacera conducida", "Engrasar chumacera lado conducido.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "COCCION", "LUB", "Lubricación cadena de transmisión", "Aplicar lubricante en cadena. Verificar tensión.", "CRC BELT GRIP", "QUINCENAL", 15),
        (tag, name, "COCCION", "INSP", "Inspección fajas y poleas", "Verificar estado, tensión y alineación de fajas. Revisar desgaste de canales de poleas.", "-", "SEMANAL", 7),
        (tag, name, "COCCION", "INSP", "Inspección motor eléctrico", "Verificar temperatura, ruido, vibración, ventilador y cableado.", "-", "QUINCENAL", 15),
    ]


def _tareas_percolador(tag, name):
    return [
        (tag, name, "COCCION", "LUB", "Lubricación chumaceras", "Engrasar chumaceras del percolador.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "COCCION", "INSP", "Inspección motor y transmisión", "Verificar motor, fajas, poleas, reductor.", "-", "SEMANAL", 7),
        (tag, name, "COCCION", "INSP", "Inspección estructura y malla", "Verificar estado de malla filtrante, soportes y estructura.", "-", "QUINCENAL", 15),
    ]


def _tareas_molino(tag, name):
    """Molinos: frecuencia INTERDIARIA (cada 2 días)"""
    return [
        (tag, name, "MOLINO", "LUB", "Lubricación chumaceras molino", "Engrasar chumaceras de eje principal del molino.", "GRASA NLGI 2", "INTERDIARIA", 2),
        (tag, name, "MOLINO", "INSP", "Inspección martillos", "Verificar desgaste, fisuras, fijación de martillos. Rotar o reemplazar según criterio.", "-", "INTERDIARIA", 2),
        (tag, name, "MOLINO", "INSP", "Inspección cribas/mallas", "Verificar estado de cribas, perforaciones, desgaste. Reemplazar si hay rotura.", "-", "INTERDIARIA", 2),
        (tag, name, "MOLINO", "INSP", "Inspección fajas y poleas", "Verificar estado, tensión, alineación de fajas. Revisar canales de poleas.", "-", "SEMANAL", 7),
        (tag, name, "MOLINO", "INSP", "Inspección motor eléctrico", "Verificar temperatura, amperaje, ruido, vibración, ventilador.", "-", "SEMANAL", 7),
        (tag, name, "MOLINO", "LUB", "Nivel aceite reductor (si aplica)", "Verificar nivel de aceite en visor de reductor.", "ACEITE ISO VG 320", "SEMANAL", 7),
    ]


def _tareas_th_molino(tag, name):
    """TH alimentadores de molino: interdiaria"""
    return [
        (tag, name, "MOLINO", "LUB", "Lubricación chumaceras", "Engrasar chumaceras del transportador.", "GRASA FRIXO 177", "INTERDIARIA", 2),
        (tag, name, "MOLINO", "LUB", "Lubricación cadena", "Aplicar lubricante en cadena de transmisión.", "CRC BELT GRIP", "SEMANAL", 7),
        (tag, name, "MOLINO", "INSP", "Inspección motor y transmisión", "Verificar motor, fajas, cadena, tensión.", "-", "SEMANAL", 7),
    ]


def _tareas_faja_transportadora(tag, name):
    return [
        (tag, name, "MOLINO", "LUB", "Lubricación rodillos y chumaceras", "Engrasar rodillos de carga, retorno y chumaceras de cabezal.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "MOLINO", "INSP", "Inspección banda transportadora", "Verificar desgaste, alineación, tensión, empalmes y bordes.", "-", "SEMANAL", 7),
        (tag, name, "MOLINO", "INSP", "Inspección motor y tambor motriz", "Verificar motor, tambor, lagging, revestimiento.", "-", "QUINCENAL", 15),
    ]


def _tareas_ciclon(tag, name, area):
    return [
        (tag, name, area, "INSP", "Inspección ductos y cono", "Verificar desgaste de ductos, cono, uniones soldadas, fugas de producto.", "-", "QUINCENAL", 15),
        (tag, name, area, "INSP", "Inspección válvula rotativa (si aplica)", "Verificar juego, desgaste de paletas, rodamientos.", "-", "QUINCENAL", 15),
    ]


def _tareas_zaranda_th(tag, name):
    return [
        (tag, name, "MOLINO", "LUB", "Lubricación chumaceras", "Engrasar chumaceras del transportador de zaranda.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "MOLINO", "INSP", "Inspección motor y transmisión", "Verificar estado general de motor, fajas, cadena.", "-", "SEMANAL", 7),
        (tag, name, "MOLINO", "INSP", "Inspección malla zaranda", "Verificar desgaste, tensión y limpieza de malla.", "-", "QUINCENAL", 15),
    ]


def _tareas_secador(tag, name):
    return [
        (tag, name, "SECADO", "LUB", "Lubricación chumaceras principales", "Engrasar chumaceras del cilindro secador.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "SECADO", "LUB", "Lubricación engranaje corona/piñón", "Aplicar grasa abierta en corona y piñón de giro.", "GRASA ABIERTA OGL", "QUINCENAL", 15),
        (tag, name, "SECADO", "LUB", "Nivel aceite reductor", "Verificar nivel de aceite en reductor de secador.", "ACEITE ISO VG 320", "QUINCENAL", 15),
        (tag, name, "SECADO", "INSP", "Inspección fajas y poleas", "Verificar estado, tensión, alineación de fajas de transmisión.", "-", "SEMANAL", 7),
        (tag, name, "SECADO", "INSP", "Inspección motor eléctrico", "Verificar temperatura, ruido, vibración, ventilador.", "-", "SEMANAL", 7),
        (tag, name, "SECADO", "INSP", "Inspección rodillos de apoyo", "Verificar desgaste, alineación y holgura de rodillos de apoyo del cilindro.", "-", "QUINCENAL", 15),
        (tag, name, "SECADO", "INSP", "Inspección sellos de entrada/salida", "Verificar desgaste de sellos, fugas de producto o aire caliente.", "-", "QUINCENAL", 15),
    ]


def _tareas_th_secado(tag, name, linea):
    return [
        (tag, name, "SECADO", "LUB", "Lubricación chumacera motriz", "Engrasar chumacera lado motriz.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "SECADO", "LUB", "Lubricación chumacera conducida", "Engrasar chumacera lado conducido.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "SECADO", "LUB", "Lubricación cadena de transmisión", "Aplicar lubricante en cadena. Verificar tensión.", "CRC BELT GRIP", "QUINCENAL", 15),
        (tag, name, "SECADO", "INSP", "Inspección fajas y poleas", "Verificar estado, tensión, alineación de fajas.", "-", "SEMANAL", 7),
        (tag, name, "SECADO", "INSP", "Inspección motor eléctrico", "Verificar temperatura, ruido, vibración.", "-", "QUINCENAL", 15),
    ]


def _tareas_purificador(tag, name):
    return [
        (tag, name, "SECADO", "LUB", "Lubricación chumaceras", "Engrasar chumaceras del purificador.", "GRASA FRIXO 177", "SEMANAL", 7),
        (tag, name, "SECADO", "INSP", "Inspección motor y transmisión", "Verificar motor, fajas, reductor.", "-", "SEMANAL", 7),
        (tag, name, "SECADO", "INSP", "Inspección mallas y tamiz", "Verificar desgaste de mallas filtrantes.", "-", "QUINCENAL", 15),
    ]


# ── Equipos y mapeo a funciones ────────────────────────────────────────────────

EQUIPOS = [
    # COCCION - Digestores D1-D9
    *[(_tareas_digestor, f"D{i}", f"DIGESTOR #{i}") for i in range(1, 10)],
    # COCCION - TH de digestores
    *[(_tareas_th_coccion, f"TH{i}", f"TH{i}") for i in range(1, 10)],
    # COCCION - Percoladores
    (_tareas_percolador, "PER1", "PERCOLADOR #1"),
    (_tareas_percolador, "PER2", "PERCOLADOR #2"),
    # MOLINO - Molinos
    (_tareas_molino, "MOLI1-LINE", "MOLINO #1"),
    (_tareas_molino, "MOLI2-LINE1", "MOLINO #2"),
    # MOLINO - TH alimentadores molino
    (_tareas_th_molino, "MOL1-TH1", "TH ALIMENTADOR MOL #1"),
    (_tareas_th_molino, "MOL2-TH1", "TH ALIMENTADOR MOL #2"),
    (_tareas_th_molino, "ZAR1-TH1", "TH ALIMENTADOR ZARANDA"),
    # MOLINO - Faja transportadora
    (_tareas_faja_transportadora, "FTRA", "FAJA TRANSPORTADORA"),
    # MOLINO - Ciclones
    ("ciclon", "ENSA1", "CICLON DE ENSAQUE"),
    ("ciclon", "LLEGA1", "CICLON DE LLEGADA"),
    # SECADO - Secadores
    (_tareas_secador, "SECA-SECA1", "SECADOR #1"),
    (_tareas_secador, "SECA-SECA2", "SECADOR #2"),
    # SECADO - Ciclones
    ("ciclon_sec", "CICL-SECA1", "CICLON DE FINOS SEC #1"),
    ("ciclon_sec", "CICL-SECA2", "CICLON DE FINOS SEC #2"),
    # SECADO - Purificador
    (_tareas_purificador, "PURI1", "PURIFICADOR"),
    # SECADO - TH Secador #1
    *[("th_sec", tag, name) for tag, name in [
        ("THAL-SECA", "TH ALIMENTADOR SEC1"), ("TH1S-SECA", "TH1 SALIDA SEC1"),
        ("TH2A-SECA", "TH2 ALIM. ENFRIADOR"), ("THFI-SECA", "TH FINO SEC1"),
        ("THSA-SECA", "TH SALIDA SEC1"), ("TH3S-ENFR", "TH3 SALIDA ENFRIADOR"),
        ("TH4S-PURI", "TH4 SALIDA PURIFICADOR"),
    ]],
    # SECADO - TH Secador #2
    *[("th_sec", f"SEC2-TH{i}", f"TH{i} SEC2") for i in range(1, 11)],
]


def gen_all_tasks():
    tasks = []
    for item in EQUIPOS:
        if item[0] == "ciclon":
            tasks.extend(_tareas_ciclon(item[1], item[2], "MOLINO"))
        elif item[0] == "ciclon_sec":
            tasks.extend(_tareas_ciclon(item[1], item[2], "SECADO"))
        elif item[0] == "th_sec":
            tasks.extend(_tareas_th_secado(item[1], item[2], "SECADOR"))
        else:
            func, tag, name = item
            tasks.extend(func(tag, name))
    return tasks


def run():
    tasks = gen_all_tasks()

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan Preventivo"

    # Estilos
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    lub_fill = PatternFill("solid", fgColor="E2EFDA")
    insp_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "N°", "AREA", "TAG", "EQUIPO", "TIPO", "ACTIVIDAD",
        "PROCEDIMIENTO", "LUBRICANTE / INSUMO", "FRECUENCIA",
        "DIAS", "TURNO", "RESPONSABLE", "CHECK", "OBSERVACIONES"
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    for i, t in enumerate(tasks, 1):
        tag, name, area, tipo, actividad, procedimiento, lubricante, frecuencia, dias = t
        fill = lub_fill if tipo == "LUB" else insp_fill
        row = [
            i, area, tag, name,
            "LUBRICACIÓN" if tipo == "LUB" else "INSPECCIÓN",
            actividad, procedimiento, lubricante,
            frecuencia, dias, "NOCHE", "PROVEEDOR", "", ""
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.fill = fill
            cell.alignment = left if cell.column in (6, 7) else center

    # Anchos de columna
    widths = {'A': 5, 'B': 12, 'C': 14, 'D': 26, 'E': 14, 'F': 36,
              'G': 55, 'H': 28, 'I': 14, 'J': 6, 'K': 10, 'L': 14, 'M': 8, 'N': 25}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'F2'
    ws.auto_filter.ref = f"A1:N{ws.max_row}"

    # ── Hoja RESUMEN ──
    ws2 = wb.create_sheet("Resumen", 0)
    ws2.append(["PLAN DE MANTENIMIENTO PREVENTIVO — TURNO NOCHE"])
    ws2['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws2.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws2.append([])
    ws2.append(["RESUMEN POR FRECUENCIA"])
    ws2['A4'].font = Font(bold=True, size=11)

    from collections import Counter
    freq_count = Counter(t[7] for t in tasks)
    tipo_count = Counter("LUBRICACIÓN" if t[3] == "LUB" else "INSPECCIÓN" for t in tasks)
    area_count = Counter(t[2] for t in tasks)

    ws2.append(["Frecuencia", "Cantidad de tareas"])
    for f in ["INTERDIARIA", "SEMANAL", "QUINCENAL"]:
        ws2.append([f, freq_count.get(f, 0)])

    ws2.append([])
    ws2.append(["RESUMEN POR TIPO"])
    ws2['A' + str(ws2.max_row)].font = Font(bold=True, size=11)
    ws2.append(["Tipo", "Cantidad"])
    for t, c in tipo_count.items():
        ws2.append([t, c])

    ws2.append([])
    ws2.append(["RESUMEN POR AREA"])
    ws2['A' + str(ws2.max_row)].font = Font(bold=True, size=11)
    ws2.append(["Área", "Cantidad"])
    for a, c in sorted(area_count.items()):
        ws2.append([a, c])

    ws2.append([])
    ws2.append(["TOTAL TAREAS", len(tasks)])
    ws2['A' + str(ws2.max_row)].font = Font(bold=True, size=12, color="FF0000")
    ws2['B' + str(ws2.max_row)].font = Font(bold=True, size=12, color="FF0000")

    ws2.append([])
    ws2.append(["NOTAS:"])
    ws2.append(["- INTERDIARIA: cada 2 días (aplica solo a molinos: martillos, cribas, chumaceras)"])
    ws2.append(["- SEMANAL: cada 7 días"])
    ws2.append(["- QUINCENAL: cada 15 días"])
    ws2.append(["- Turno: NOCHE — ejecutado por proveedor de servicios"])
    ws2.append(["- Verde = Lubricación, Azul = Inspección"])
    ws2.append(["- Cualquier hallazgo reportar como Aviso en el CMMS"])

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'plan_preventivo_proveedor.xlsx')
    wb.save(out)
    print(f"Excel generado: {out}")
    print(f"Total tareas: {len(tasks)}")
    for f in ["INTERDIARIA", "SEMANAL", "QUINCENAL"]:
        print(f"  {f}: {freq_count.get(f, 0)}")
    for t, c in tipo_count.items():
        print(f"  {t}: {c}")


if __name__ == '__main__':
    run()
